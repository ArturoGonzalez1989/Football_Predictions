"""
telegram_bt_crossref.py
=======================
Cruza las señales de Telegram (ChatExport/messages.html) con las apuestas BT
del último bt_results_*.csv y genera un Excel de análisis en analisis/.

Lógica de matching:
  1. Extrae strategy_label y match_name de cada señal Telegram
     - Formato nuevo: "🟢 STRATEGY_LABEL TEAM_A - TEAM_B · BET @ ODDS · Min MIN | SCORE"
     - Formato antiguo: "SEÑAL DETECTADA ⚽ MATCH 📊 STRATEGY 💰 BET @ ODDS ⏱ Min MIN | Score: SCORE"
  2. Normaliza nombres de partido a slug sin acentos
  3. Busca en BT: mismo strategy_label, solapamiento de nombre >= 0.4, ±3 min
  4. Filtra BT al rango de fechas de las señales (+/-6h)

Output Excel — 3 hojas:
  - SIGNALS_vs_BT   : señal Telegram + columnas BT si hay match
  - BT_ONLY         : apuestas BT del período sin señal Telegram
  - STATS           : resumen cuantitativo
"""

import re
import unicodedata
import glob
import sys
import io
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import urlparse, unquote
from bs4 import BeautifulSoup
import pandas as pd

# Fix Windows console encoding
if sys.stdout.encoding.lower() != "utf-8":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

# ── Rutas ─────────────────────────────────────────────────────────────────────
PROJECT_ROOT = Path(__file__).resolve().parent.parent
CHAT_HTML    = PROJECT_ROOT / "ChatExport" / "messages.html"
ANALISIS_DIR = PROJECT_ROOT / "analisis"

# ── Normalización ─────────────────────────────────────────────────────────────

def _remove_accents(s: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )

def _slugify(s: str) -> str:
    """'Kocaelispor - Konyaspor' → 'kocaelispor konyaspor'"""
    s = _remove_accents(s.lower())
    s = re.sub(r"[^\w\s]", " ", s)
    return re.sub(r"\s+", " ", s).strip()

def _slug_from_match_id(match_id: str) -> str:
    """'kocaelispor-konyaspor-apuestas-35350900' → 'kocaelispor konyaspor'"""
    parts = match_id.rsplit("-apuestas-", 1)
    name = parts[0] if parts else match_id
    name = _remove_accents(name.lower())
    return re.sub(r"[-_]+", " ", name).strip()

def _name_overlap(slug_a: str, slug_b: str) -> float:
    wa = set(slug_a.split())
    wb = set(slug_b.split())
    if not wa or not wb:
        return 0.0
    return len(wa & wb) / len(wa | wb)

# ── Cargador BT ───────────────────────────────────────────────────────────────

def _load_latest_bt(analisis_dir: Path):
    csvs = sorted(glob.glob(str(analisis_dir / "bt_results_*.csv")))
    if not csvs:
        raise FileNotFoundError(f"No se encontró bt_results_*.csv en {analisis_dir}")
    latest = csvs[-1]
    print(f"[BT] Cargando: {Path(latest).name}")
    df = pd.read_csv(latest)
    df["timestamp_utc"] = pd.to_datetime(df["timestamp_utc"], errors="coerce")
    df["bt_slug"] = df["match_id"].apply(_slug_from_match_id)
    return df, Path(latest).name

# ── Parser de señales Telegram ────────────────────────────────────────────────

def _strip_emojis_start(s: str) -> str:
    """Elimina emojis y espacios del inicio del string."""
    return re.sub(r"^[\U0001F300-\U0001FFFE\u2600-\u26FF\u2700-\u27BF\s]+", "", s)

def _parse_new_format(text: str, known_labels: list[str]) -> dict | None:
    """
    Formato: 🟢 STRATEGY_LABEL TEAM_A - TEAM_B · BET @ ODDS | mín: X · Min MINUTE' | SCORE | …

    Estrategia: busca el label conocido más largo que coincida al inicio del texto.
    Si no, usa heurística: todo antes del primer ' - ' que sea seguido de letras de equipo.
    """
    clean = _strip_emojis_start(text)

    # Separar antes del primer ' · '
    parts = re.split(r"\s+·\s+", clean, maxsplit=1)
    if len(parts) < 2:
        return None
    prefix  = parts[0].strip()   # "STRATEGY_LABEL TEAM_A - TEAM_B"
    rest    = parts[1]            # "BET @ ODDS | mín: X · Min MINUTE' | SCORE | …"

    # Alias: variantes de nombres usadas en Telegram vs. BT
    LABEL_ALIASES = {
        "xg underperformance (base)"    : "xG Underperformance",
        "xg underperformance base"      : "xG Underperformance",
        "back draw xg convergence"      : "BACK Draw xG Conv",
        "back u3.5 late"                : "BACK Under 3.5 Late",
        "back u4.5 3 goals"             : "BACK Under 4.5 3 Goals",
        "back o2.5 2-goal lead"         : "BACK O2.5 2-Goal Lead",
        "back o2.5 two goals"           : "BACK O2.5 Two Goals",
        "back over 0.5 poss extreme"    : "BACK Over 0.5 Poss Extreme",
        "lay over 4.5 blowout"          : "LAY Over 4.5 Blowout",
        "lay over 4.5 v3"               : "LAY Over 4.5 v3",
    }

    # --- Extraer strategy_label y match_name del prefix ---
    # Intento 1: busca label conocido al inicio (longest match first)
    # Requiere que el label sea seguido de espacio o fin de string (word boundary)
    strategy_label = None
    match_name     = None
    for label in known_labels:
        label_norm = re.sub(r"\s+", " ", label.strip())
        pref_low   = prefix.lower()
        lab_low    = label_norm.lower()
        if pref_low.startswith(lab_low):
            # Verificar word boundary: el siguiente char debe ser espacio o fin
            next_pos = len(lab_low)
            if next_pos >= len(pref_low) or pref_low[next_pos] == " ":
                strategy_label = label_norm
                match_name     = prefix[next_pos:].strip()
                break

    # Intento 1b: busca alias conocidos
    if strategy_label is None:
        for alias_low, canonical in LABEL_ALIASES.items():
            if prefix.lower().startswith(alias_low):
                next_pos = len(alias_low)
                if next_pos >= len(prefix) or prefix[next_pos] == " ":
                    strategy_label = canonical
                    match_name     = prefix[next_pos:].strip()
                    break

    # Limpiar match_name de sufijos residuales como "(Base)"
    if match_name:
        match_name = re.sub(r"^\([^)]+\)\s*", "", match_name).strip()

    # Intento 2: el partido es la parte TEAM_A - TEAM_B al final del prefix
    #   Asume que el primer ' - ' visible que no forma parte de una cuota/score
    #   divide strategy de match.
    if strategy_label is None:
        # Busca el primer ' - ' rodeado de palabras (no números)
        m = re.search(r"(?<=[A-Za-záéíóúüñÁÉÍÓÚÜÑ])\s+-\s+(?=[A-Za-záéíóúüñÁÉÍÓÚÜÑ])", prefix)
        if m:
            # Determina el punto de corte: busca hacia atrás el inicio del equipo local
            # (suele ser 1-3 palabras antes del ' - ')
            before_dash = prefix[:m.start()]
            words = before_dash.split()
            # El nombre del equipo local son las últimas 1-3 palabras
            # La estrategia son las palabras anteriores
            # Heurística: el equipo local empieza en la última palabra en mayúscula seguida de minúscula
            team_start = max(0, len(words) - 3)
            for i in range(len(words) - 1, -1, -1):
                if words[i][0].isupper() and (i == 0 or not words[i-1][0].islower()):
                    team_start = i
                    break
            strategy_label = " ".join(words[:team_start]).strip()
            match_name     = prefix[prefix.index(words[team_start]) if team_start < len(words) else 0:].strip()

    if not strategy_label:
        strategy_label = ""
    if not match_name:
        match_name = prefix

    # --- Extraer odds, minuto, score del resto ---
    odds_m   = re.search(r"@\s*([\d.]+)", rest)
    minute_m = re.search(r"Min\s+(\d+)", rest)
    score_m  = re.search(r"\|\s*(\d+-\d+)\s*\|", rest)
    if not score_m:
        score_m = re.search(r"(\d+-\d+)", rest)

    if not minute_m:
        return None

    return {
        "strategy_label": strategy_label,
        "match"         : match_name,
        "bet_desc"      : rest.split("@")[0].strip() if "@" in rest else "",
        "odds"          : float(odds_m.group(1)) if odds_m else 0.0,
        "minute"        : int(minute_m.group(1)),
        "score"         : score_m.group(1) if score_m else "",
    }

def _parse_old_format(text: str) -> dict | None:
    """
    Formato: 🔴 SEÑAL DETECTADA ⚽ MATCH 📊 STRATEGY 💰 BET @ ODDS @ ODDS2 ⏱ Min MIN' | Score: SCORE
    """
    m = re.search(
        r"⚽\s+(.+?)\s+📊\s+(.+?)\s+💰\s+(.+?)@\s*([\d.]+).*?Min\s+(\d+).*?Score:\s*(\d+-\d+)",
        text, re.DOTALL
    )
    if not m:
        return None
    return {
        "strategy_label": m.group(2).strip(),
        "match"         : m.group(1).strip(),
        "bet_desc"      : m.group(3).strip(),
        "odds"          : float(m.group(4)),
        "minute"        : int(m.group(5)),
        "score"         : m.group(6),
    }

def _parse_telegram_html(html_path: Path, known_labels: list[str]) -> list[dict]:
    """Extrae todas las señales del bot del HTML exportado de Telegram."""
    with open(html_path, encoding="utf-8") as f:
        soup = BeautifulSoup(f.read(), "html.parser")

    # Labels ordenados por longitud desc para que el match sea greedy
    labels_sorted = sorted(known_labels, key=len, reverse=True)

    records = []
    current_date = None
    current_sender = None   # rastrea quién habla en mensajes "joined"
    signal_color = {"🔴": "RED", "🟢": "GREEN", "🟡": "YELLOW"}

    for div in soup.find_all("div", class_="message"):
        classes = div.get("class", [])

        # Separador de fecha
        if "service" in classes:
            raw = div.get_text(strip=True)
            for fmt in ("%d %B %Y", "%d %b %Y"):
                try:
                    current_date = datetime.strptime(raw, fmt)
                    break
                except ValueError:
                    pass
            continue

        # Actualizar sender cuando hay from_name explícito
        fn = div.find("div", class_="from_name")
        if fn:
            current_sender = "furbo" if "furbo" in fn.text.lower() else "user"

        # Solo procesar mensajes del bot (explícitos o joined)
        if current_sender != "furbo":
            continue

        text_div = div.find("div", class_="text")
        if not text_div:
            continue
        text = text_div.get_text(separator=" ", strip=True)

        # Extraer match_id desde el primer href que contenga "-apuestas-"
        tg_match_id = None
        for a in text_div.find_all("a", href=True):
            href = unquote(a["href"])
            seg = href.rstrip("/").split("/")[-1]
            if "-apuestas-" in seg:
                tg_match_id = seg
                break

        # Timestamp exacto
        date_div = div.find("div", class_="date details")
        msg_ts = None
        if date_div and date_div.get("title"):
            raw_ts = re.sub(r"\s+UTC[+\-]\d+:\d+$", "", date_div["title"])
            try:
                msg_ts = datetime.strptime(raw_ts, "%d.%m.%Y %H:%M:%S")
            except ValueError:
                pass
        if msg_ts is None and current_date:
            msg_ts = current_date

        # Color de señal
        color = "UNKNOWN"
        for emoji, name in signal_color.items():
            if emoji in text[:3]:
                color = name
                break

        # Parsear según formato
        if "SEÑAL DETECTADA" in text:
            rec = _parse_old_format(text)
        else:
            rec = _parse_new_format(text, labels_sorted)

        if rec is None:
            continue

        records.append({
            "tg_timestamp"     : msg_ts,
            "tg_color"         : color,
            "tg_match"         : rec["match"].strip(),
            "tg_match_id"      : tg_match_id,
            "tg_strategy_label": rec["strategy_label"].strip(),
            "tg_bet_desc"      : rec["bet_desc"].strip(),
            "tg_odds"          : rec["odds"],
            "tg_minute"        : rec["minute"],
            "tg_score"         : rec["score"],
            "tg_raw"           : text[:250],
        })

    return records

# ── Matching ──────────────────────────────────────────────────────────────────
OVERLAP_THRESHOLD  = 0.35  # mínimo solapamiento de palabras en nombre del partido

# Aliases: variantes de nombre de estrategia en Telegram → label en BT
LABEL_ALIASES = {
    "xg underperformance (base)"   : "xg underperformance",
    "xg underperformance base"     : "xg underperformance",
    "back draw xg convergence"     : "back draw xg conv",
    "back u3.5 late"               : "back under 3.5 late",
    "back u3.5 3-goal lid"         : "back under 3.5 3 goals",
    "back u4.5 3-goals low xg"     : "back under 4.5 3 goals",
    "back o2.5 2-goal lead"        : "back over 2.5 2-goal lead",
    "back o2.5 two goals"          : "back over 2.5 two goals",
    "back over 0.5 poss extreme"   : "back poss extreme",
    "lay over 4.5 blowout"         : "lay over 4.5 blowout",
    "lay over 4.5 v3"              : "lay over 4.5 v3",
    "back draw 0-0 (v2r)"          : "back draw 0-0",
    "goal clustering"              : "goal clustering",
    "pressure cooker"              : "pressure cooker",
}

def _normalize_label(label: str) -> str:
    """Normaliza un label de estrategia aplicando aliases conocidos."""
    low = label.strip().lower()
    return LABEL_ALIASES.get(low, low)

def _labels_match(tg_label: str, bt_label: str) -> bool:
    """True si los labels coinciden (exacto o fuzzy ≥2 palabras)."""
    a = _normalize_label(tg_label)
    b = _normalize_label(bt_label)
    if a == b:
        return True
    wa = set(a.split())
    wb = set(b.split())
    return len(wa & wb) >= 2

def _find_bt_match(tg: dict, bt_df: pd.DataFrame):
    """
    Devuelve (fila_bt, calidad_match, overlap_score).

    Objetivo: localizar el resultado BT (won/pl) para calcular rendimiento real
    de las señales Telegram. El minuto se usa solo como desempate, no como filtro
    de exclusión — lo que importa es match_id + estrategia.

    Prioridad:
    1. match_id exacto (de URL) + label exacto
    2. match_id exacto + label fuzzy
    3. match_id exacto + cualquier estrategia del mismo partido (MATCH_ONLY)
    4. Fallback fuzzy por nombre + label
    """
    tg_label    = tg["tg_strategy_label"].strip()
    tg_min      = tg["tg_minute"]
    tg_ts       = tg["tg_timestamp"]
    tg_match_id = tg.get("tg_match_id")

    # ── Filtro temporal ±24h ──────────────────────────────────────────────────
    if tg_ts is not None:
        lo = tg_ts - timedelta(hours=24)
        hi = tg_ts + timedelta(hours=24)
        subset = bt_df[
            bt_df["timestamp_utc"].notna() &
            (bt_df["timestamp_utc"].dt.tz_localize(None) >= lo) &
            (bt_df["timestamp_utc"].dt.tz_localize(None) <= hi)
        ].copy()
    else:
        subset = bt_df.copy()

    if subset.empty:
        return None, "NO_DATE_MATCH", 0.0

    # ── match_id exacto desde URL ─────────────────────────────────────────────
    if tg_match_id:
        by_id = subset[subset["match_id"] == tg_match_id].copy()
        if not by_id.empty:
            # Intento 1: label exacto o alias
            exact = by_id[by_id["strategy_label"].apply(
                lambda s: _labels_match(tg_label, s)
            )]
            if not exact.empty:
                # Desempate: el más cercano en minuto
                exact = exact.copy()
                exact["_dm"] = (exact["minuto"] - tg_min).abs()
                best = exact.sort_values("_dm").iloc[0]
                q = "EXACT" if best["minuto"] == tg_min else "GOOD"
                return best, q, 1.0

            # Intento 2: partido correcto pero estrategia no matchea en BT
            return None, "NO_STRATEGY_MATCH", 1.0

    # ── Fallback: fuzzy por nombre + label ───────────────────────────────────
    tg_slug = _slugify(tg["tg_match"])
    subset = subset.copy()
    subset["_ov"] = subset["bt_slug"].apply(lambda s: _name_overlap(tg_slug, s))
    subset = subset[subset["_ov"] >= OVERLAP_THRESHOLD]
    if subset.empty:
        return None, "NO_NAME_MATCH", 0.0

    matched = subset[subset["strategy_label"].apply(
        lambda s: _labels_match(tg_label, s)
    )]
    if matched.empty:
        return None, "NO_STRATEGY_MATCH", 0.0

    best = matched.sort_values("_ov", ascending=False).iloc[0]
    q = "EXACT" if best["_ov"] >= 0.5 and best["minuto"] == tg_min else "GOOD"
    return best, q, float(best["_ov"])

# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("[1] Cargando BT results...")
    bt_df, bt_filename = _load_latest_bt(ANALISIS_DIR)
    known_labels = sorted(bt_df["strategy_label"].dropna().unique().tolist(), key=len, reverse=True)
    print(f"    -> {len(bt_df)} apuestas BT | {len(known_labels)} estrategias unicas")

    print("[2] Parseando senales Telegram...")
    signals = _parse_telegram_html(CHAT_HTML, known_labels)
    print(f"    -> {len(signals)} senales parseadas")
    if signals:
        for s in signals[:3]:
            print(f"       {s['tg_strategy_label']!r:35} | {s['tg_match']!r:35} | min={s['tg_minute']} odds={s['tg_odds']}")

    # Rango de fechas: desde la PRIMERA señal Telegram (dinámico) hasta la última +1d
    tg_dates = [s["tg_timestamp"] for s in signals if s["tg_timestamp"]]
    lo_date  = min(tg_dates) if tg_dates else datetime(2026, 3, 14)
    hi_date  = max(tg_dates) + timedelta(days=1) if tg_dates else datetime(2026, 3, 17)

    # Primera señal para trazabilidad
    first_sig = min(signals, key=lambda s: s["tg_timestamp"] or datetime.max) if signals else None
    if first_sig:
        print(f"    -> Primera senal: {first_sig['tg_timestamp']} | {first_sig['tg_match']} | {first_sig['tg_strategy_label']}")

    # BT en el rango: solo partidos desde la primera señal en adelante
    bt_period = bt_df[
        bt_df["timestamp_utc"].notna() &
        (bt_df["timestamp_utc"].dt.tz_localize(None) >= lo_date) &
        (bt_df["timestamp_utc"].dt.tz_localize(None) <= hi_date)
    ].copy()
    bt_period_idx = set(bt_period.index)
    print(f"    -> {len(bt_period)} apuestas BT en el periodo ({lo_date.strftime('%Y-%m-%d %H:%M')} - {hi_date.date()})")

    print("[3] Cruzando senales con BT...")
    rows_crossref = []
    matched_bt_indices = set()

    for sig in signals:
        bt_row, quality, overlap = _find_bt_match(sig, bt_period)

        row = {
            "tg_timestamp"      : sig["tg_timestamp"],
            "tg_color"          : sig["tg_color"],
            "tg_match"          : sig["tg_match"],
            "tg_match_id"       : sig.get("tg_match_id"),
            "tg_strategy"       : sig["tg_strategy_label"],
            "tg_bet"            : sig["tg_bet_desc"],
            "tg_odds"           : sig["tg_odds"],
            "tg_minute"         : sig["tg_minute"],
            "tg_score"          : sig["tg_score"],
            "match_quality"     : quality,
            "name_overlap"      : round(overlap, 2),
            "bt_match_id"       : None,
            "bt_strategy_key"   : None,
            "bt_strategy_label" : None,
            "bt_minute"         : None,
            "bt_odds"           : None,
            "bt_won"            : None,
            "bt_pl"             : None,
            "bt_country"        : None,
            "bt_league"         : None,
            "bt_timestamp"      : None,
            "diff_min"          : None,
            "diff_odds"         : None,
        }

        if bt_row is not None:
            idx = bt_row.name
            matched_bt_indices.add(idx)
            row.update({
                "bt_match_id"       : bt_row["match_id"],
                "bt_strategy_key"   : bt_row["strategy"],
                "bt_strategy_label" : bt_row["strategy_label"],
                "bt_minute"         : int(bt_row["minuto"]),
                "bt_odds"           : float(bt_row["back_odds"]),
                "bt_won"            : bt_row["won"],
                "bt_pl"             : float(bt_row["pl"]),
                "bt_country"        : bt_row.get("Pais", bt_row.get("País", "")),
                "bt_league"         : bt_row.get("Liga", ""),
                "bt_timestamp"      : bt_row["timestamp_utc"],
                "diff_min"          : int(sig["tg_minute"]) - int(bt_row["minuto"]),
                "diff_odds"         : round(sig["tg_odds"] - float(bt_row["back_odds"]), 2),
            })

        rows_crossref.append(row)

    df_cross = pd.DataFrame(rows_crossref)

    # Separar matcheadas (config actual) de descartadas (config anterior)
    MATCHED_QUALITIES = {"EXACT", "GOOD", "FUZZY"}
    df_matched   = df_cross[df_cross["match_quality"].isin(MATCHED_QUALITIES)].copy()
    df_discarded = df_cross[~df_cross["match_quality"].isin(MATCHED_QUALITIES)].copy()

    # BT_ONLY: apuestas BT del periodo sin señal Telegram
    bt_only_idx = bt_period_idx - matched_bt_indices
    df_bt_only  = bt_period.loc[list(bt_only_idx)].drop(
        columns=[c for c in ["bt_slug","_ov","_lm","_fl"] if c in bt_period.columns],
        errors="ignore"
    ).sort_values("timestamp_utc", na_position="last")

    # ── Stats ─────────────────────────────────────────────────────────────────
    n_signals   = len(signals)
    n_exact     = (df_matched["match_quality"] == "EXACT").sum()
    n_good      = (df_matched["match_quality"] == "GOOD").sum()
    n_fuzzy     = (df_matched["match_quality"] == "FUZZY").sum()
    n_matched   = len(df_matched)
    n_unmatched = len(df_discarded)
    n_bt_only   = len(df_bt_only)

    matched_rows = df_matched[df_matched["bt_pl"].notna()].copy()
    total_pl     = matched_rows["bt_pl"].sum()
    won_count    = (matched_rows["bt_won"] == True).sum()
    lost_count   = (matched_rows["bt_won"] == False).sum()
    n_resolved   = won_count + lost_count
    wr_bt        = won_count / n_resolved * 100 if n_resolved > 0 else 0
    roi_bt       = total_pl / n_resolved * 100 if n_resolved > 0 else 0

    # Rendimiento por estrategia
    by_strat = (
        matched_rows.groupby("tg_strategy")
        .agg(
            N        = ("bt_pl", "count"),
            Won      = ("bt_won", lambda x: (x == True).sum()),
            PL       = ("bt_pl", "sum"),
        )
        .assign(
            WR_pct   = lambda d: (d["Won"] / d["N"] * 100).round(1),
            ROI_pct  = lambda d: (d["PL"] / d["N"] * 100).round(1),
        )
        .sort_values("PL", ascending=False)
        .reset_index()
    )
    by_strat.columns = ["Estrategia", "N", "Ganadas", "P/L", "WR%", "ROI%"]

    stats_rows = [
        {"Metrica": "Senales Telegram totales",          "Valor": n_signals},
        {"Metrica": "Matcheadas con resultado BT",       "Valor": n_matched},
        {"Metrica": "  · Calidad EXACT",                 "Valor": int(n_exact)},
        {"Metrica": "  · Calidad GOOD",                  "Valor": int(n_good)},
        {"Metrica": "Descartadas (config anterior)",       "Valor": n_unmatched},
        {"Metrica": "Apuestas BT sin senal Telegram",    "Valor": n_bt_only},
        {"Metrica": "Match rate (%)",                    "Valor": round(n_matched / n_signals * 100, 1) if n_signals else 0},
        {"Metrica": "── Rendimiento señales Telegram ──","Valor": ""},
        {"Metrica": "N con resultado conocido",          "Valor": int(n_resolved)},
        {"Metrica": "Ganadas",                           "Valor": int(won_count)},
        {"Metrica": "Perdidas",                          "Valor": int(lost_count)},
        {"Metrica": "WR%",                               "Valor": round(wr_bt, 1)},
        {"Metrica": "P/L total (£1 stake c/u)",         "Valor": round(total_pl, 2)},
        {"Metrica": "ROI%",                              "Valor": round(roi_bt, 1)},
        {"Metrica": "Fichero BT usado",                  "Valor": bt_filename},
        {"Metrica": "Generado",                          "Valor": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
    ]
    stats = pd.DataFrame(stats_rows)

    # ── Excel ─────────────────────────────────────────────────────────────────
    ts_str   = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path = ANALISIS_DIR / f"telegram_bt_crossref_{ts_str}.xlsx"

    print(f"[4] Exportando Excel -> {out_path.name}")
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df_matched.to_excel(writer, sheet_name="SIGNALS_vs_BT", index=False)
        df_bt_only.to_excel(writer, sheet_name="BT_ONLY", index=False)
        df_discarded.to_excel(writer, sheet_name="DESCARTADAS_config_anterior", index=False)
        stats.to_excel(writer, sheet_name="STATS", index=False)
        by_strat.to_excel(writer, sheet_name="POR_ESTRATEGIA", index=False)

        from openpyxl.styles import PatternFill, Font
        COLORS = {
            "EXACT"              : "C6EFCE",
            "GOOD"               : "FFEB9C",
            "FUZZY"              : "FFD966",
            "NO_DATE_MATCH"      : "FFC7CE",
            "NO_NAME_MATCH"      : "FFC7CE",
            "NO_MINUTE_MATCH"    : "FFC7CE",
            "NO_STRATEGY_MATCH"  : "FFC7CE",
            "NONE"               : "FFC7CE",
        }
        ws = writer.sheets["SIGNALS_vs_BT"]
        col_names = [c.value for c in ws[1]]
        mq_col = (col_names.index("match_quality") + 1) if "match_quality" in col_names else None
        if mq_col:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                cell = row[mq_col - 1]
                fill_color = COLORS.get(str(cell.value), "FFFFFF")
                cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)

        for sheet_name in writer.sheets:
            ws2 = writer.sheets[sheet_name]
            for cell in ws2[1]:
                cell.font = Font(bold=True)
            ws2.freeze_panes = "A2"

    print(f"\nListo: {out_path}")
    print(f"  Senales: {n_signals} | Matcheadas: {n_matched} ({round(n_matched/n_signals*100,1) if n_signals else 0}%) | Sin match: {n_unmatched}")
    print(f"  ── Rendimiento señales Telegram (N={n_resolved}) ──")
    print(f"  WR: {round(wr_bt,1)}% | P/L: {round(total_pl,2)} | ROI: {round(roi_bt,1)}%  (£1 stake por señal)")
    print(f"  Ganadas: {won_count} | Perdidas: {lost_count}")
    print(f"\n  ── Por estrategia ──")
    print(by_strat.to_string(index=False))


if __name__ == "__main__":
    main()
