#!/usr/bin/env python3
"""
Generate static HTML analytics dashboard from the cartera API.

Usage:  python analisis/generate_dashboard.py
Output: analisis/dashboard_portfolio.html
Requires: backend running at localhost:8000
"""

import sys, json, math, statistics, glob
from datetime import datetime, timezone
from collections import defaultdict
from pathlib import Path

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8")

try:
    import requests
except ImportError:
    print("ERROR: 'requests' no instalado. Ejecuta: pip install requests")
    sys.exit(1)

# ── Config ──────────────────────────────────────────────────────────────

BASE_DIR = Path(__file__).resolve().parent.parent
API_URL = "http://localhost:8000/api/analytics/strategies/cartera"
CONFIG_PATH = BASE_DIR / "betfair_scraper" / "cartera_config.json"
DATA_DIR = BASE_DIR / "betfair_scraper" / "data"
OUTPUT_PATH = Path(__file__).resolve().parent / "dashboard_bt.html"
LIVE_OUTPUT_PATH = Path(__file__).resolve().parent / "dashboard_live.html"
TEAM_LOOKUP_PATH = Path(__file__).resolve().parent / "team_lookup.json"
FLAT_STAKE = 1


# ── Data fetching ───────────────────────────────────────────────────────

def _build_match_meta():
    """Build match_id → {País, Liga, ft_score} dict from all partido_*.csv files."""
    import csv as _csv
    import re as _re
    from urllib.parse import unquote as _unquote

    match_meta = {}
    for pcsv in DATA_DIR.glob("partido_*.csv"):
        try:
            with open(pcsv, encoding="utf-8") as f:
                rows = list(_csv.DictReader(f))
            if not rows:
                continue
            r0, rl = rows[0], rows[-1]
            mid = r0.get("tab_id") or pcsv.stem.replace("partido_", "")
            pais = r0.get("País", "")
            liga = r0.get("Liga", "")
            if (not pais or not liga) and r0.get("url"):
                url_parts = r0["url"].split("/")
                if len(url_parts) >= 7:
                    liga = liga or _unquote(url_parts[-2]).replace("-", " ").title()
            entry = {"País": pais, "Liga": liga, "ft_score": rl.get("score", "")}
            for key in [mid, _unquote(mid)]:
                if key and (key not in match_meta or (not match_meta[key]["País"] and pais)):
                    match_meta[key] = entry
        except Exception:
            continue

    _PAIS_FIXES = {
        "thun-lausanne-apuestas-35244305": ("Suiza", "Super League"),
    }
    for _mid, (_pais, _liga) in _PAIS_FIXES.items():
        if _mid in match_meta and not match_meta[_mid]["País"]:
            match_meta[_mid]["País"] = _pais
            match_meta[_mid]["Liga"] = match_meta[_mid]["Liga"] or _liga

    return match_meta


def _enrich_bet(b, match_meta):
    """Attach País, Liga, ft_score to a bet dict from match_meta."""
    import re as _re
    mid = b["match_id"]
    meta = match_meta.get(mid) or match_meta.get(_re.sub(r"-DESKTOP-\w+$", "", mid), {})
    b["País"]     = meta.get("País", "")
    b["Liga"]     = meta.get("Liga", "")
    b["ft_score"] = meta.get("ft_score", "")
    return b


def fetch_bt_bets():
    """Read the most recent portfolio_bets_*.xlsx (or .csv) exported by the notebook."""
    import csv as _csv

    bt_dir = Path(__file__).resolve().parent
    candidates = sorted(bt_dir.glob("portfolio_bets_*.xlsx"), reverse=True)
    use_excel = bool(candidates)
    if not candidates:
        candidates = sorted(bt_dir.glob("portfolio_bets_*.csv"), reverse=True)
    if not candidates:
        print("ERROR: No se encontró portfolio_bets_*.xlsx ni .csv en analisis/")
        print("Ejecuta el notebook strategies_designer.ipynb (Kernel → Restart & Run All).")
        sys.exit(1)

    bt_csv = candidates[0]
    print(f"  Leyendo BT {'Excel' if use_excel else 'CSV'}: {bt_csv.name}")
    match_meta = _build_match_meta()

    if use_excel:
        import openpyxl as _xl
        wb = _xl.load_workbook(bt_csv, read_only=True, data_only=True)
        ws = wb.active
        headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        rows_iter = (
            {headers[i]: (c.value if c.value is not None else "") for i, c in enumerate(row)}
            for row in ws.iter_rows(min_row=2)
        )
    else:
        rows_iter = _csv.DictReader(open(bt_csv, encoding="utf-8"))

    bets = []
    for row in rows_iter:
        pl_raw = float(row.get("pl_eur") or row.get("pl") or 0)
        won_raw = row.get("won", "")
        odds = float(row.get("effective_odds") or row.get("back_odds") or 0) or None
        b = {
            "match_id":      row.get("match_id", ""),
            "strategy":      row.get("strategy", ""),
            "timestamp_utc": row.get("timestamp_utc", ""),
            "minuto":        int(float(row["minuto"])) if row.get("minuto") else None,
            "won":           won_raw in ("True", "1", "true", True),
            "pl":            round(pl_raw * FLAT_STAKE, 2),
            "back_odds":     odds,
            "team":          row.get("team") or row.get("backed_team") or "",
            "risk_level":    row.get("risk_level", ""),
        }
        bets.append(_enrich_bet(b, match_meta))

    print(f"  BT bets leídas: {len(bets)}")
    return bets


def fetch_live_bets():
    """Read placed_bets.csv (paper trading) and normalize to the same bet dict format."""
    import csv as _csv

    csv_path = BASE_DIR / "betfair_scraper" / "placed_bets.csv"
    if not csv_path.exists():
        print(f"  AVISO: {csv_path} no encontrado, LIVE vacío")
        return []

    match_meta = _build_match_meta()

    bets = []
    with open(csv_path, encoding="utf-8") as f:
        for row in _csv.DictReader(f):
            if row.get("status", "") == "pending":
                continue  # skip unresolved bets
            pl_raw = float(row.get("pl") or 0)
            odds = float(row.get("back_odds") or 0) or None
            b = {
                "match_id":      row.get("match_id", ""),
                "strategy":      row.get("strategy", ""),
                "timestamp_utc": row.get("timestamp_utc", ""),
                "minuto":        int(float(row["minute"])) if row.get("minute") else None,
                "won":           row.get("result", "") == "won",
                "pl":            round(pl_raw, 2),
                "back_odds":     odds,
                "team":          "",
                "risk_level":    row.get("confidence", ""),
            }
            bets.append(_enrich_bet(b, match_meta))

    print(f"  LIVE bets leídas: {len(bets)}")
    return bets


def fetch_data():
    print("Fetching data from API...")
    try:
        r = requests.get(API_URL, timeout=120)
        r.raise_for_status()
    except Exception as e:
        print(f"ERROR: No se pudo conectar al backend: {e}")
        print("Asegúrate de que el backend está corriendo en localhost:8000")
        sys.exit(1)
    data = r.json()
    print(f"  Recibidas {len(data.get('bets', []))} bets crudas")
    return data


def load_config():
    with open(CONFIG_PATH, encoding="utf-8") as f:
        return json.load(f)


def count_matches():
    return len(list(DATA_DIR.glob("partido_*.csv")))


def fetch_sd_bets(cfg):
    """Generate SD strategy bets directly from CSV files via aux/sd_generators."""
    import sys
    aux_dir = str(BASE_DIR / "aux")
    if aux_dir not in sys.path:
        sys.path.insert(0, aux_dir)
    try:
        from sd_generators import generate_all_new_bets
    except ImportError as e:
        print(f"  WARNING: No se pudo importar sd_generators: {e}")
        return []

    strats = cfg.get("strategies", {})
    enabled_sd = {k for k, v in strats.items()
                  if k.startswith("sd_") and isinstance(v, dict) and v.get("enabled")}
    if not enabled_sd:
        return []

    print(f"  Generando bets SD ({len(enabled_sd)} estrategias activas)...")
    all_sd = generate_all_new_bets(str(DATA_DIR))
    print(f"  SD brutas: {len(all_sd)}")

    result = []
    for bet in all_sd:
        s = bet.get("strategy", "")
        if s not in enabled_sd:
            continue

        # Apply config minute range
        sc = strats.get(s, {})
        m = bet.get("minuto")
        if m is not None and (m < sc.get("m_min", 0) or m > sc.get("m_max", 90)):
            continue

        # Normalize odds to back_odds (generators use various field names)
        if not bet.get("back_odds"):
            for odds_key in (
                "back_cs_odds", "back_cs_00_odds",
                "back_ud_odds",
                "back_home_fav_odds", "back_home_odds",
                "back_away_odds",
                "back_longshot_odds",
                "back_over_odds", "back_over25_odds", "back_over05_odds",
                "back_under35_odds", "back_under45_odds",
                "back_draw_11_odds", "back_draw_conv_odds",
                "lay_over45_odds",
                "over_odds",
            ):
                if bet.get(odds_key):
                    bet["back_odds"] = bet[odds_key]
                    break

        # Scale P/L from stake=1 (generators) to FLAT_STAKE=10 (dashboard)
        bet["pl"] = round(bet["pl"] * FLAT_STAKE, 2)

        # Ensure compatibility with apply_realistic_adjustments:
        # - stability_count=99 so SD bets always pass the stability filter
        # - team=strategy so dedup uses (match_id, strategy) as key instead of (match_id, Unknown)
        bet["stability_count"] = 99
        bet.setdefault("team", bet.get("strategy", "sd"))

        result.append(bet)

    print(f"  SD tras filtros: {len(result)}")
    return result


# ── Filtering (mirrors cartera.ts) ──────────────────────────────────────

def get_bet_odds(b):
    return b.get("back_draw") or b.get("back_over_odds") or b.get("over_odds") or b.get("back_odds") or 2.0


def get_bet_type(b):
    strat = b.get("strategy", "")
    if strat == "back_draw_00":
        return "Draw"
    ol = b.get("over_line", "")
    if ol:
        return ol.replace("Over ", "O ")
    return b.get("team", "Unknown")


def bet_market_key(b):
    return f"{b.get('match_id', '')}::{get_bet_type(b)}"


def filter_by_strategy(bets, cfg):
    """Filter bets keeping only enabled strategies with matching config params.
    Mirrors cartera.ts filter*Bets() functions."""
    strats = cfg.get("strategies", {})
    result = []
    for b in bets:
        s = b.get("strategy", "")

        if s == "back_draw_00":
            sc = strats.get("draw", {})
            if not sc.get("enabled", True):
                continue
            xg_max = sc.get("xgMax", 1.0)
            poss_max = sc.get("possMax", 100)
            shots_max = sc.get("shotsMax", 20)
            m_min = sc.get("minuteMin", 0)
            m_max = sc.get("minuteMax", 90)
            if xg_max < 1.0 and b.get("xg_total") is not None and b["xg_total"] >= xg_max:
                continue
            if poss_max < 100 and b.get("poss_diff") is not None and b["poss_diff"] >= poss_max:
                continue
            if shots_max < 20 and b.get("shots_total") is not None and b["shots_total"] >= shots_max:
                continue
            if m_min > 0 and b.get("minuto") is not None and b["minuto"] < m_min:
                continue
            if m_max < 90 and b.get("minuto") is not None and b["minuto"] >= m_max:
                continue

        elif s == "xg_underperformance":
            sc = strats.get("xg", {})
            if not sc.get("enabled", True):
                continue
            sot_min = sc.get("sotMin", 0)
            xg_excess_min = sc.get("xgExcessMin", 0)
            m_min = sc.get("minuteMin", 0)
            m_max = sc.get("minuteMax", 90)
            if xg_excess_min > 0 and b.get("xg_excess") is not None and b["xg_excess"] < xg_excess_min:
                continue
            if sot_min > 0 and b.get("sot_team") is not None and b["sot_team"] < sot_min:
                continue
            if m_min > 0 and b.get("minuto") is not None and b["minuto"] < m_min:
                continue
            if m_max < 90 and b.get("minuto") is not None and b["minuto"] >= m_max:
                continue

        elif s == "odds_drift":
            sc = strats.get("drift", {})
            if not sc.get("enabled", True):
                continue
            goal_diff_min = sc.get("goalDiffMin", 0)
            drift_min = sc.get("driftMin", 30)
            odds_max = sc.get("oddsMax", 999)
            m_min = sc.get("minuteMin", 0)
            m_max = sc.get("minuteMax", 90)
            mom_gap_min = sc.get("momGapMin", 0)
            if goal_diff_min > 0 and b.get("goal_diff") is not None and b["goal_diff"] < goal_diff_min:
                continue
            if drift_min > 30 and b.get("drift_pct") is not None and b["drift_pct"] < drift_min:
                continue
            if odds_max < 999 and b.get("back_odds") is not None and b["back_odds"] > odds_max:
                continue
            if m_min > 0 and b.get("minuto") is not None and b["minuto"] < m_min:
                continue
            if m_max < 90 and b.get("minuto") is not None and b["minuto"] >= m_max:
                continue
            if mom_gap_min > 0 and b.get("synth_momentum_gap") is not None and b["synth_momentum_gap"] <= mom_gap_min:
                continue

        elif s == "goal_clustering":
            sc = strats.get("clustering", {})
            if not sc.get("enabled", True):
                continue
            sot_min = sc.get("sotMin", 0)
            m_min = sc.get("minuteMin", 0)
            m_max = sc.get("minuteMax", 90)
            xg_rem_min = sc.get("xgRemMin", 0)
            if sot_min > 0 and b.get("sot_max") is not None and b["sot_max"] < sot_min:
                continue
            if m_min > 0 and b.get("minuto") is not None and b["minuto"] < m_min:
                continue
            if m_max < 90 and b.get("minuto") is not None and b["minuto"] >= m_max:
                continue
            if xg_rem_min > 0 and b.get("synth_xg_remaining") is not None and b["synth_xg_remaining"] < xg_rem_min:
                continue

        elif s == "pressure_cooker":
            sc = strats.get("pressure", {})
            if not sc.get("enabled", True):
                continue
            m_min = sc.get("minuteMin", 0)
            m_max = sc.get("minuteMax", 90)
            if m_min > 0 and b.get("minuto") is not None and b["minuto"] < m_min:
                continue
            if m_max < 90 and b.get("minuto") is not None and b["minuto"] >= m_max:
                continue

        elif s == "tarde_asia":
            sc = strats.get("tarde_asia", {})
            if not sc.get("enabled", True):
                continue

        elif s in ("momentum_xg_v1", "momentum_xg_v2"):
            sc = strats.get("momentum_xg", {})
            ver = sc.get("version", "v1")
            if ver == "off":
                continue
            if s == "momentum_xg_v1" and ver != "v1":
                continue
            if s == "momentum_xg_v2" and ver != "v2":
                continue
            m_min = sc.get("minuteMin", 0)
            m_max = sc.get("minuteMax", 90)
            if m_min > 0 and b.get("minuto") is not None and b["minuto"] < m_min:
                continue
            if m_max < 90 and b.get("minuto") is not None and b["minuto"] >= m_max:
                continue

        result.append(b)
    return result


def apply_realistic_adjustments(bets, cfg):
    """Mirror cartera.ts:applyRealisticAdjustments + filterByRisk."""
    adj = cfg.get("adjustments", {})
    if not adj.get("enabled", True):
        return bets

    result = list(bets)
    global_min = adj.get("global_minute_min")
    global_max = adj.get("global_minute_max")
    drift_min_min = adj.get("drift_min_minute", 0)
    max_odds = adj.get("max_odds", 999)
    min_odds = adj.get("min_odds", 0)
    dedup = adj.get("dedup", False)
    conflict_filter = adj.get("conflict_filter", False)
    allow_contrarias = adj.get("allow_contrarias", True)
    stability = adj.get("stability", 1)
    slippage_pct = adj.get("slippage_pct", 0)

    # 0. Global minute range
    if global_min is not None or global_max is not None:
        filtered = []
        for b in result:
            m = b.get("minuto")
            if m is None:
                filtered.append(b)
                continue
            if global_min is not None and m < global_min:
                continue
            if global_max is not None and m >= global_max:
                continue
            filtered.append(b)
        result = filtered

    # 1. Drift min minute
    if drift_min_min > 0:
        result = [b for b in result
                  if b.get("strategy") != "odds_drift"
                  or (b.get("minuto") is not None and b["minuto"] >= drift_min_min)]

    # 2. Max odds
    if max_odds < 999:
        result = [b for b in result if get_bet_odds(b) <= max_odds]

    # 3. Min odds
    if min_odds > 0:
        result = [b for b in result if get_bet_odds(b) >= min_odds]

    # 4. Dedup
    if dedup:
        seen = set()
        deduped = []
        for b in result:
            key = bet_market_key(b)
            if key not in seen:
                seen.add(key)
                deduped.append(b)
        result = deduped

    # 5. Conflict filter
    if conflict_filter:
        xg_matches = {b["match_id"] for b in result if b.get("strategy") == "xg_underperformance"}
        result = [b for b in result
                  if b.get("strategy") not in ("momentum_xg_v1", "momentum_xg_v2")
                  or b.get("match_id") not in xg_matches]

    # 5b. Anti-contrarias
    if not allow_contrarias:
        seen_mo = {}
        filtered = []
        for b in result:
            is_mo = b.get("strategy") in ("back_draw_00", "odds_drift", "momentum_xg_v1", "momentum_xg_v2")
            if not is_mo:
                filtered.append(b)
                continue
            bt = "draw" if b.get("strategy") == "back_draw_00" else (b.get("team") or "home")
            mid = b.get("match_id")
            first = seen_mo.get(mid)
            if first is None:
                seen_mo[mid] = bt
                filtered.append(b)
            elif first == bt:
                filtered.append(b)
            # else: contraria, skip
        result = filtered

    # 6. Stability
    if stability > 1:
        result = [b for b in result if (b.get("stability_count") or 1) >= stability]

    # 7. Slippage (modify P/L for wins)
    if slippage_pct > 0:
        factor = 1 - slippage_pct / 100
        adjusted = []
        for b in result:
            if not b.get("won"):
                adjusted.append(b)
            else:
                nb = dict(b)
                odds = get_bet_odds(b) * factor
                nb["pl"] = round((odds - 1) * FLAT_STAKE * 0.95, 2)
                adjusted.append(nb)
        result = adjusted

    # Risk filter
    risk_filter = cfg.get("risk_filter", "all")
    if risk_filter != "all":
        filtered = []
        for b in result:
            rl = (b.get("risk_info") or {}).get("risk_level", "none")
            if risk_filter == "no_risk" and rl not in ("none", ""):
                continue
            if risk_filter == "medium" and rl != "medium":
                continue
            if risk_filter == "high" and rl != "high":
                continue
            if risk_filter == "with_risk" and rl not in ("medium", "high"):
                continue
            filtered.append(b)
        result = filtered

    return result


# ── Metrics computation ─────────────────────────────────────────────────

def compute_kpis(bets, n_matches):
    n = len(bets)
    if n == 0:
        return {"n_matches": n_matches, "n_bets": 0, "wr": 0, "roi": 0, "pl": 0,
                "max_dd": 0, "best_streak": 0, "worst_streak": 0, "sharpe": 0,
                "profit_factor": 0, "cumulative": []}

    wins = sum(1 for b in bets if b.get("won"))
    pls = [b.get("pl", 0) for b in bets]
    total_pl = sum(pls)
    total_staked = n * FLAT_STAKE

    # Cumulative P/L
    cum = []
    running = 0
    for p in pls:
        running += p
        cum.append(round(running, 2))

    # Max drawdown
    peak = 0
    max_dd = 0
    for v in cum:
        if v > peak:
            peak = v
        dd = peak - v
        if dd > max_dd:
            max_dd = dd

    # Streaks
    best_w = worst_l = cur_w = cur_l = 0
    for b in bets:
        if b.get("won"):
            cur_w += 1
            cur_l = 0
            best_w = max(best_w, cur_w)
        else:
            cur_l += 1
            cur_w = 0
            worst_l = max(worst_l, cur_l)

    # Sharpe
    mean_pl = statistics.mean(pls)
    std_pl = statistics.stdev(pls) if n > 1 else 1
    sharpe = round(mean_pl / std_pl * math.sqrt(n), 2) if std_pl > 0 else 0

    # Profit factor
    gross_win = sum(p for p in pls if p > 0)
    gross_loss = abs(sum(p for p in pls if p < 0))
    pf = round(gross_win / gross_loss, 2) if gross_loss > 0 else float("inf")

    return {
        "n_matches": n_matches,
        "n_bets": n,
        "wr": round(wins / n * 100, 1),
        "roi": round(total_pl / total_staked * 100, 1),
        "pl": round(total_pl, 2),
        "max_dd": round(max_dd, 2),
        "best_streak": best_w,
        "worst_streak": worst_l,
        "sharpe": sharpe,
        "profit_factor": pf,
        "cumulative": cum,
    }


def _group_stats(bets, key_fn):
    groups = defaultdict(list)
    for b in bets:
        k = key_fn(b)
        if k:
            groups[k].append(b)
    table = []
    for k, gb in groups.items():
        n = len(gb)
        w = sum(1 for b in gb if b.get("won"))
        pl = sum(b.get("pl", 0) for b in gb)
        table.append({
            "label": k, "n": n, "wins": w,
            "wr": round(w / n * 100, 1) if n else 0,
            "roi": round(pl / (n * FLAT_STAKE) * 100, 1) if n else 0,
            "pl": round(pl, 2),
        })
    table.sort(key=lambda x: x["n"], reverse=True)
    return table


STRATEGY_LABELS = {
    "back_draw_00": "Back Empate 0-0",
    "xg_underperformance": "xG Underperf",
    "odds_drift": "Odds Drift",
    "goal_clustering": "Goal Clustering",
    "pressure_cooker": "Pressure Cooker",
    "tarde_asia": "Tarde Asia",
    "momentum_xg_v1": "Momentum xG v1",
    "momentum_xg_v2": "Momentum xG v2",
    # SD strategies
    "sd_over25_2goal": "SD Over 2.5 (2-goal lead)",
    "sd_under35_late": "SD Under 3.5 Late",
    "sd_longshot": "SD Longshot",
    "sd_cs_close": "SD CS Close (2-1/1-2)",
    "sd_cs_one_goal": "SD CS One-Goal (1-0/0-1)",
    "sd_ud_leading": "SD Underdog Leading",
    "sd_home_fav_leading": "SD Home Fav Leading",
    "sd_cs_20": "SD CS 2-0/0-2",
    "sd_cs_big_lead": "SD CS Big Lead",
}


def compute_strategy_table(bets):
    return _group_stats(bets, lambda b: STRATEGY_LABELS.get(b.get("strategy"), b.get("strategy", "?")))


def compute_geo_table(bets, field):
    table = _group_stats(bets, lambda b: b.get(field) or "Desconocido")
    # Push "Desconocido/a" to end
    known = [r for r in table if r["label"] not in ("Desconocido", "Desconocida")]
    unknown = [r for r in table if r["label"] in ("Desconocido", "Desconocida")]
    return (known[:19] + unknown)[:20]


def compute_day_of_week(bets):
    days = ["Lun", "Mar", "Mié", "Jue", "Vie", "Sáb", "Dom"]
    groups = defaultdict(list)
    for b in bets:
        ts = b.get("timestamp_utc", "")
        if not ts:
            continue
        try:
            dt = datetime.fromisoformat(ts.replace("Z", "+00:00"))
            groups[dt.weekday()].append(b)
        except Exception:
            pass
    result = []
    for i, day in enumerate(days):
        gb = groups.get(i, [])
        n = len(gb)
        w = sum(1 for x in gb if x.get("won"))
        pl = sum(x.get("pl", 0) for x in gb)
        result.append({
            "label": day, "n": n,
            "wr": round(w / n * 100, 1) if n else 0,
            "roi": round(pl / (n * FLAT_STAKE) * 100, 1) if n else 0,
            "pl": round(pl, 2),
        })
    return result


def compute_hour_of_day(bets):
    groups = defaultdict(list)
    for b in bets:
        ts = b.get("timestamp_utc", "")
        if not ts:
            continue
        try:
            dt = datetime.fromisoformat(ts.replace("Z", "+00:00"))
            groups[dt.hour].append(b)
        except Exception:
            pass
    result = []
    for h in range(24):
        gb = groups.get(h, [])
        n = len(gb)
        w = sum(1 for x in gb if x.get("won"))
        pl = sum(x.get("pl", 0) for x in gb)
        result.append({
            "label": f"{h:02d}h", "n": n,
            "wr": round(w / n * 100, 1) if n else 0,
            "pl": round(pl, 2),
        })
    return result


def compute_monthly_pl(bets):
    groups = defaultdict(list)
    for b in bets:
        ts = b.get("timestamp_utc", "")
        if not ts:
            continue
        try:
            dt = datetime.fromisoformat(ts.replace("Z", "+00:00"))
            groups[dt.strftime("%Y-%m")].append(b)
        except Exception:
            pass
    months = sorted(groups.keys())
    result = []
    for m in months:
        gb = groups[m]
        pl = sum(x.get("pl", 0) for x in gb)
        n = len(gb)
        w = sum(1 for x in gb if x.get("won"))
        result.append({
            "label": m, "pl": round(pl, 2), "n": n,
            "wr": round(w / n * 100, 1) if n else 0,
        })
    return result


def compute_minute_hist(bets):
    groups = defaultdict(list)
    for b in bets:
        m = b.get("minuto")
        if m is None:
            continue
        bucket = (int(m) // 5) * 5
        groups[bucket].append(b)
    result = []
    for bucket in range(0, 95, 5):
        gb = groups.get(bucket, [])
        n = len(gb)
        w = sum(1 for x in gb if x.get("won"))
        result.append({
            "label": f"{bucket}-{bucket+5}",
            "n": n,
            "wr": round(w / n * 100, 1) if n else 0,
        })
    return result


LEAGUE_TIERS = {
    # Tier 1 — Top 5 European leagues
    "Premier League": 1, "La Liga": 1, "Bundesliga": 1,
    "Serie A": 1, "Ligue 1": 1,
    # Tier 2 — Secondary European, top international, established leagues
    "Championship": 2, "Segunda División": 2, "2. Bundesliga": 2,
    "Serie B": 2, "Ligue 2": 2,
    "Eredivisie": 2, "Primeira Liga": 2, "Jupiler Pro League": 2,
    "Süper Lig": 2, "Super League": 2, "Superligaen": 2,
    "Premiership": 2,  # Scotland
    "Liga Argentina": 2, "Série A": 2, "Liga MX": 2, "MLS": 2,
    "J-League": 2, "Saudi Pro League": 2, "K1 League": 2,
    "Champions League": 2, "Europa League": 2, "Conference League": 2,
    "Copa Libertadores": 2, "Copa Sudamericana": 2,
    "FA Cup": 2, "Copa del Rey": 2, "Coppa Italia": 2, "DFB-Pokal": 2,
    "Coupe de France": 2, "Copa Turca": 2,
    "Liga 1": 2,  # Romania
    "Liga Checa": 2,
    "AFC Champions League": 2,
}
TIER_LABELS = {1: "Top 5 Ligas", 2: "Ligas Secundarias", 3: "Ligas Menores"}
TIER_COLORS = {1: "#3b82f6", 2: "#f59e0b", 3: "#8b5cf6"}

# Tournament type classification
TOURNAMENT_TYPES = {
    # Liga doméstica — Top
    "Liga Top": [
        "Premier League", "La Liga", "Bundesliga", "Serie A", "Ligue 1",
    ],
    # Liga doméstica — Segunda/Tercera
    "Liga 2ª/3ª": [
        "Championship", "League One", "League Two", "National League",
        "Segunda División", "Liga de Expansión",
        "2. Bundesliga", "Serie B", "Serie C",
        "Ligue 2", "Eerste Divisie", "Premiership",
        "I Liga", "Prva Liga", "Liga A", "Liga 2",
        "Liga Checa", "Superligaen",
        "1st Division", "Premier Division",
    ],
    # Liga internacional (fuera de Europa)
    "Liga Internacional": [
        "Liga Argentina", "Série A", "Baiano", "Carioca", "Gaúcho",
        "Goiano", "Mineiro", "Paulista Serie A1", "Pernambucano",
        "Copa do Brasil",   # BRA domestic cup
        "Liga MX", "MLS", "Liga Colombiana", "Liga Uruguaya",
        "J-League", "K1 League", "Saudi Pro League", "Iran Pro League",
        "Qatar Stars League", "UAE Pro League", "Virsliga",
        "Super League", "Superliga", "Jupiler Pro League",
        "Primeira Liga", "Eredivisie", "Süper Lig",
        "AFC Champions League 2",
        "Primera División",   # varios países latinoamericanos
        "Super League Suiza", "Liga 1",  # Suiza / Rumanía / Perú
    ],
    # Copa nacional
    "Copa Nacional": [
        "FA Cup", "Challenge Cup", "Copa Escocesa",
        "Copa del Rey", "Copa Turca",
        "Coppa Italia", "DFB-Pokal",
        "Coupe de France",
        "Copa Argentina",
    ],
    # Competición europea / continental
    "Competición Continental": [
        "Champions League", "Europa League", "Conference League",
        "Copa Libertadores", "Copa Sudamericana",
        "AFC Champions League", "CONCACAF Champions",
    ],
    # Selecciones / amistosos
    "Selecciones / Amistosos": [
        "FIFA Women World Cup", "FIFA Women WC Qualifiers",
        "Torneo Amistoso",
    ],
}
# Invert: liga → type
_LIGA_TO_TYPE = {}
for _ttype, _ligas in TOURNAMENT_TYPES.items():
    for _l in _ligas:
        _LIGA_TO_TYPE[_l] = _ttype
_TOURNAMENT_ORDER = list(TOURNAMENT_TYPES.keys()) + ["Otros"]
_TOURNAMENT_COLORS = {
    "Liga Top": "#3b82f6",
    "Liga 2ª/3ª": "#6366f1",
    "Liga Internacional": "#f59e0b",
    "Copa Nacional": "#ec4899",
    "Competición Continental": "#10b981",
    "Selecciones / Amistosos": "#8b5cf6",
    "Otros": "#6b7280",
}


def get_league_tier(liga):
    return LEAGUE_TIERS.get(liga, 3)


def get_tournament_type(liga):
    return _LIGA_TO_TYPE.get(liga, "Otros")


def compute_tournament_type_table(bets):
    groups = defaultdict(list)
    for b in bets:
        liga = b.get("Liga") or "Desconocida"
        ttype = get_tournament_type(liga)
        groups[ttype].append(b)
    result = []
    for ttype in _TOURNAMENT_ORDER:
        gb = groups.get(ttype, [])
        if not gb:
            continue
        n = len(gb)
        w = sum(1 for x in gb if x.get("won"))
        pl = sum(x.get("pl", 0) for x in gb)
        result.append({
            "label": ttype,
            "n": n,
            "wr": round(w / n * 100, 1) if n else 0,
            "roi": round(pl / (n * FLAT_STAKE) * 100, 1) if n else 0,
            "pl": round(pl, 2),
        })
    return result


def compute_league_tier_table(bets):
    groups = defaultdict(list)
    for b in bets:
        liga = b.get("Liga") or "Desconocida"
        tier = get_league_tier(liga)
        groups[tier].append(b)
    result = []
    for tier in [1, 2, 3]:
        gb = groups.get(tier, [])
        n = len(gb)
        w = sum(1 for x in gb if x.get("won"))
        pl = sum(x.get("pl", 0) for x in gb)
        result.append({
            "label": TIER_LABELS[tier],
            "tier": tier,
            "n": n,
            "wr": round(w / n * 100, 1) if n else 0,
            "roi": round(pl / (n * FLAT_STAKE) * 100, 1) if n else 0,
            "pl": round(pl, 2),
        })
    return result


def load_team_lookup():
    """Load match_id → [home, away] mapping."""
    if TEAM_LOOKUP_PATH.exists():
        with open(TEAM_LOOKUP_PATH, encoding="utf-8") as f:
            return json.load(f)
    return {}


def _parse_ft_result(ft_score):
    """Parse ft_score like '2-1' into result category: 'home_win', 'draw', 'away_win', or None."""
    if not ft_score or not isinstance(ft_score, str):
        return None
    parts = ft_score.strip().split("-")
    if len(parts) != 2:
        return None
    try:
        h, a = int(parts[0].strip()), int(parts[1].strip())
    except ValueError:
        return None
    if h > a:
        return "home_win"
    elif h == a:
        return "draw"
    else:
        return "away_win"


def compute_team_yield(bets, team_lookup):
    """Compute profitability by team involvement.
    Each bet is attributed to both teams playing in the match.
    Returns ALL teams (no min filter — JS handles that).
    Includes País, Liga, and ft_score result breakdowns.
    """
    from collections import Counter

    teams = defaultdict(lambda: {
        "n": 0, "wins": 0, "pl": 0,
        "n_home": 0, "pl_home": 0, "wins_home": 0,
        "n_away": 0, "pl_away": 0, "wins_away": 0,
        "n_draw": 0, "pl_draw": 0,
        "n_home_win": 0, "pl_home_win": 0,
        "n_away_win": 0, "pl_away_win": 0,
        "paises": Counter(), "ligas": Counter(),
    })

    for b in bets:
        mid = b.get("match_id", "")
        lookup = team_lookup.get(mid)
        if not lookup or len(lookup) < 2:
            continue
        home_team, away_team = lookup[0], lookup[1]
        if not home_team or not away_team:
            continue

        won = b.get("won", False)
        pl = b.get("pl", 0)
        pais = b.get("País") or ""
        liga = b.get("Liga") or ""
        ft_result = _parse_ft_result(b.get("ft_score"))

        for team_name, is_home in [(home_team, True), (away_team, False)]:
            t = teams[team_name]
            t["n"] += 1
            t["pl"] += pl
            t["wins"] += 1 if won else 0
            if pais:
                t["paises"][pais] += 1
            if liga:
                t["ligas"][liga] += 1

            if is_home:
                t["n_home"] += 1
                t["pl_home"] += pl
                t["wins_home"] += 1 if won else 0
            else:
                t["n_away"] += 1
                t["pl_away"] += pl
                t["wins_away"] += 1 if won else 0

            if ft_result == "draw":
                t["n_draw"] += 1
                t["pl_draw"] += pl
            elif ft_result == "home_win":
                t["n_home_win"] += 1
                t["pl_home_win"] += pl
            elif ft_result == "away_win":
                t["n_away_win"] += 1
                t["pl_away_win"] += pl

    table = []
    for name, t in teams.items():
        pais = t["paises"].most_common(1)[0][0] if t["paises"] else ""
        liga = t["ligas"].most_common(1)[0][0] if t["ligas"] else ""
        table.append({
            "label": name, "pais": pais, "liga": liga,
            "n": t["n"], "wins": t["wins"],
            "wr": round(t["wins"] / t["n"] * 100, 1) if t["n"] else 0,
            "roi": round(t["pl"] / (t["n"] * FLAT_STAKE) * 100, 1) if t["n"] else 0,
            "pl": round(t["pl"], 2),
            "n_home": t["n_home"],
            "pl_home": round(t["pl_home"], 2),
            "wr_home": round(t["wins_home"] / t["n_home"] * 100, 1) if t["n_home"] else 0,
            "n_away": t["n_away"],
            "pl_away": round(t["pl_away"], 2),
            "wr_away": round(t["wins_away"] / t["n_away"] * 100, 1) if t["n_away"] else 0,
            "n_draw": t["n_draw"],
            "pl_draw": round(t["pl_draw"], 2),
            "n_home_win": t["n_home_win"],
            "pl_home_win": round(t["pl_home_win"], 2),
            "n_away_win": t["n_away_win"],
            "pl_away_win": round(t["pl_away_win"], 2),
        })
    return table


def compute_odds_hist(bets):
    bins = [(1.0, 1.5), (1.5, 2.0), (2.0, 2.5), (2.5, 3.0), (3.0, 4.0), (4.0, 5.0), (5.0, 999)]
    labels = ["1.0-1.5", "1.5-2.0", "2.0-2.5", "2.5-3.0", "3.0-4.0", "4.0-5.0", "5.0+"]
    groups = defaultdict(list)
    for b in bets:
        odds = get_bet_odds(b)
        for i, (lo, hi) in enumerate(bins):
            if lo <= odds < hi:
                groups[i].append(b)
                break
    result = []
    for i, label in enumerate(labels):
        gb = groups.get(i, [])
        n = len(gb)
        w = sum(1 for x in gb if x.get("won"))
        result.append({
            "label": label, "n": n,
            "wr": round(w / n * 100, 1) if n else 0,
        })
    return result


# ── HTML Template ───────────────────────────────────────────────────────

def render_html(data, mode="bt"):
    data_json = json.dumps(data, ensure_ascii=False)
    kpis = data["kpis"]

    def kpi_card(label, value, fmt="", color=""):
        # Parse numeric from formatted strings like "+420.52"
        num = value if isinstance(value, (int, float)) else None
        if num is None and isinstance(value, str):
            try:
                num = float(value.replace("+", ""))
            except (ValueError, AttributeError):
                pass
        c = color or ("#10b981" if num is not None and num > 0 else "#ef4444" if num is not None and num < 0 else "#f3f4f6")
        return f'''<div class="card">
          <div class="kpi-label">{label}</div>
          <div class="kpi-value" style="color:{c}">{value}{fmt}</div>
        </div>'''

    _tbl_idx = [0]
    def table_html(rows, title):
        if not rows:
            return f"<h3>{title}</h3><p>Sin datos</p>"
        _tbl_idx[0] += 1
        tid = f"tbl{_tbl_idx[0]}"
        cols = [("", "str"), ("N", "num"), ("WR%", "num"), ("ROI%", "num"), ("P/L €", "num")]
        html = f'<h3>{title}</h3><table id="{tid}"><thead><tr>'
        for i, (col, _) in enumerate(cols):
            if col:
                html += f'<th onclick="sortTable(\'{tid}\',{i})" style="cursor:pointer;user-select:none">{col} <span class="sort-arrow"></span></th>'
            else:
                html += f"<th>{col}</th>"
        html += "</tr></thead><tbody>"
        for r in rows:
            wr_c = "#10b981" if r["wr"] >= 60 else "#f59e0b" if r["wr"] >= 50 else "#ef4444"
            pl_c = "#10b981" if r["pl"] > 0 else "#ef4444" if r["pl"] < 0 else "#9ca3af"
            roi_c = "#10b981" if r["roi"] > 0 else "#ef4444" if r["roi"] < 0 else "#9ca3af"
            html += f'''<tr>
              <td data-val="{r["label"]}">{r["label"]}</td>
              <td data-val="{r["n"]}">{r["n"]}</td>
              <td data-val="{r["wr"]}" style="color:{wr_c}">{r["wr"]}%</td>
              <td data-val="{r["roi"]}" style="color:{roi_c}">{r["roi"]}%</td>
              <td data-val="{r["pl"]}" style="color:{pl_c}">{r["pl"]:+.2f}</td>
            </tr>'''
        html += "</tbody></table>"
        return html

    kpi_cards = "".join([
        kpi_card("Total Partidos", kpis["n_matches"], color="#3b82f6"),
        kpi_card("Total Apuestas", kpis["n_bets"], color="#3b82f6"),
        kpi_card("Win Rate", kpis["wr"], "%"),
        kpi_card("ROI", kpis["roi"], "%"),
        kpi_card("Flat P/L", f'{kpis["pl"]:+.2f}', " €"),
        kpi_card("Max Drawdown", f'{kpis["max_dd"]:.2f}', " €", "#ef4444"),
        kpi_card("Mejor Racha", kpis["best_streak"], " W", "#10b981"),
        kpi_card("Peor Racha", kpis["worst_streak"], " L", "#ef4444"),
        kpi_card("Sharpe Ratio", kpis["sharpe"]),
        kpi_card("Profit Factor", kpis["profit_factor"]),
    ])

    strategy_html = table_html(data["strategy_table"], "Rendimiento por Estrategia")

    # Strategy P/L horizontal bar chart
    strat_names = json.dumps([s["label"] for s in data["strategy_table"]], ensure_ascii=False)
    strat_pls = json.dumps([s["pl"] for s in data["strategy_table"]])
    strat_colors = json.dumps(["#10b981" if s["pl"] >= 0 else "#ef4444" for s in data["strategy_table"]])
    strat_chart_height = max(180, len(data["strategy_table"]) * 32)
    country_html = table_html(data["country_table"], "Distribución por País")
    league_html = table_html(data["league_table"], "Distribución por Liga")
    dow_html = table_html(data["day_of_week"], "Día de la Semana")
    torneo_html = table_html(data["tournament_type"], "Distribución por Tipo de Torneo")

    # League tiers section
    tiers = data.get("league_tiers", [])
    tier_table = table_html(tiers, "Rendimiento por Nivel de Liga")
    tier_labels = json.dumps([t["label"] for t in tiers], ensure_ascii=False)
    tier_pls = json.dumps([t["pl"] for t in tiers])
    tier_ns = json.dumps([t["n"] for t in tiers])
    tier_wrs = json.dumps([t["wr"] for t in tiers])
    tier_chart_colors = json.dumps([TIER_COLORS.get(t.get("tier", 3), "#8b5cf6") for t in tiers])

    # Team yield summary (top 5 + worst 5 for the portfolio page)
    team_yield = data.get("team_yield", [])
    by_pl_filtered = sorted([t for t in team_yield if t["n"] >= 3], key=lambda t: t["pl"], reverse=True)
    top5 = by_pl_filtered[:5]
    worst5 = list(reversed(by_pl_filtered[-5:])) if len(by_pl_filtered) > 5 else []
    team_summary_rows = ""
    for r in (top5 + worst5):
        pl_c = "#10b981" if r["pl"] > 0 else "#ef4444" if r["pl"] < 0 else "#9ca3af"
        team_summary_rows += f'<tr><td>{r["label"]}</td><td>{r["n"]}</td><td style="color:{pl_c}">{r["pl"]:+.2f}</td></tr>'
    team_summary_html = f'''<h3>Top/Peor Equipos (resumen)</h3>
    <table><thead><tr><th>Equipo</th><th>N</th><th>P/L €</th></tr></thead><tbody>{team_summary_rows}</tbody></table>
    '''

    adj = data.get("config_summary", {}).get("adjustments", {})
    filters_info = f'dedup={adj.get("dedup")}, odds=[{adj.get("min_odds")}-{adj.get("max_odds")}], slippage={adj.get("slippage_pct")}%, stability={adj.get("stability")}, risk={data.get("config_summary", {}).get("risk_filter", "all")}'

    return f'''<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>{'BackTest' if mode == 'bt' else 'LIVE'} — Betfair Dashboard</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<style>
* {{ margin:0; padding:0; box-sizing:border-box; }}
body {{ background:#111827; color:#f3f4f6; font-family:system-ui,-apple-system,sans-serif; padding:1.5rem; }}
h1 {{ font-size:1.5rem; margin-bottom:0.25rem; }}
h2 {{ font-size:1.1rem; color:#9ca3af; margin-bottom:1.5rem; font-weight:400; }}
h3 {{ font-size:0.95rem; color:#d1d5db; margin-bottom:0.75rem; }}
.kpi-grid {{ display:grid; grid-template-columns:repeat(auto-fill,minmax(160px,1fr)); gap:0.75rem; margin-bottom:1.5rem; }}
.card {{ background:#1a2233; border:1px solid #1f2937; border-radius:8px; padding:1rem; }}
.kpi-label {{ font-size:0.7rem; color:#6b7280; text-transform:uppercase; letter-spacing:0.05em; margin-bottom:0.25rem; }}
.kpi-value {{ font-size:1.5rem; font-weight:700; }}
.chart-card {{ background:#1a2233; border:1px solid #1f2937; border-radius:8px; padding:1rem; margin-bottom:1.5rem; }}
.row {{ display:grid; grid-template-columns:1fr 1fr; gap:1.5rem; margin-bottom:1.5rem; }}
@media(max-width:768px) {{ .row {{ grid-template-columns:1fr; }} }}
table {{ width:100%; border-collapse:collapse; font-size:0.85rem; }}
th {{ color:#6b7280; font-size:0.7rem; text-transform:uppercase; border-bottom:1px solid #374151; padding:6px 8px; text-align:left; }}
td {{ padding:6px 8px; border-bottom:1px solid #1f2937; }}
tr:hover {{ background:#1f2937; }}
.footer {{ margin-top:2rem; padding-top:1rem; border-top:1px solid #1f2937; color:#6b7280; font-size:0.75rem; }}
canvas {{ max-height:300px; }}
.nav {{ display:flex; gap:0.5rem; margin-bottom:1rem; }}
.nav a {{ color:#d1d5db; text-decoration:none; padding:0.4rem 1rem; border-radius:6px; background:#1f2937; font-size:0.85rem; }}
.nav a.active {{ background:#3b82f6; color:#fff; }}
.nav a:hover {{ background:#374151; }}
</style>
</head>
<body>

<div class="nav">
  <a href="dashboard_bt.html" {'class="active"' if mode == 'bt' else ''}>BackTest</a>
  <a href="dashboard_live.html" {'class="active"' if mode == 'live' else ''}>LIVE</a>
</div>
<h1>{'BackTest' if mode == 'bt' else 'LIVE — Paper Trading'}</h1>
<h2>Generado: {data["generated_at"][:16].replace("T"," ")} — {'Filtros: ' + filters_info if mode == 'bt' else 'Apuestas paper desde placed_bets.csv'}</h2>

<div class="kpi-grid">{kpi_cards}</div>

<div class="chart-card">
  <h3>P/L Acumulado (€)</h3>
  <canvas id="cumChart"></canvas>
</div>

<div class="row">
  <div class="card">{strategy_html}
    <div style="margin-top:1rem;">
      <canvas id="stratChart" style="min-height:{strat_chart_height}px;max-height:{strat_chart_height}px;"></canvas>
    </div>
  </div>
  <div class="chart-card">
    <h3>P/L Mensual (€)</h3>
    <canvas id="monthlyChart"></canvas>
  </div>
</div>

<div class="row">
  <div class="card">{country_html}</div>
  <div class="card">{league_html}</div>
</div>

<div class="row">
  <div class="card">{torneo_html}</div>
  <div class="card"></div>
</div>

<div class="row">
  <div class="card">{tier_table}
    <div style="margin-top:1rem;">
      <canvas id="tierPLChart" style="max-height:200px;"></canvas>
    </div>
  </div>
  <div class="chart-card">
    <h3>Distribución de Apuestas por Nivel</h3>
    <canvas id="tierDonutChart" style="max-height:260px;"></canvas>
  </div>
</div>

<div class="row">
  <div class="card">{team_summary_html}</div>
  <div class="card"></div>
</div>

<div class="row">
  <div class="chart-card">
    <h3>Distribución por Minuto de Trigger</h3>
    <canvas id="minuteChart"></canvas>
  </div>
  <div class="chart-card">
    <h3>Distribución de Odds</h3>
    <canvas id="oddsChart"></canvas>
  </div>
</div>

<div class="row">
  <div class="card">{dow_html}
    <div style="margin-top:1rem;">
      <canvas id="dowChart" style="max-height:200px;"></canvas>
    </div>
  </div>
  <div class="chart-card">
    <h3>P/L por Hora del Día (€)</h3>
    <canvas id="hourChart"></canvas>
  </div>
</div>

<div class="footer">
  Dashboard de cartera Betfair — {kpis["n_matches"]} partidos analizados, {kpis["n_bets"]} apuestas filtradas
</div>

<script>
const D = {data_json};

const chartDefaults = {{
  color: '#9ca3af',
  borderColor: '#374151',
  font: {{ family: 'system-ui' }}
}};
Chart.defaults.color = '#9ca3af';
Chart.defaults.borderColor = '#1f2937';

// 1. Cumulative P/L
new Chart(document.getElementById('cumChart'), {{
  type: 'line',
  data: {{
    labels: D.cum_labels,
    datasets: [{{
      data: D.kpis.cumulative,
      borderColor: '#10b981',
      backgroundColor: 'rgba(16,185,129,0.08)',
      fill: true, tension: 0.3, pointRadius: 0, borderWidth: 2,
    }}]
  }},
  options: {{
    plugins: {{
      legend: {{ display: false }},
      annotation: undefined,
    }},
    scales: {{
      x: {{
        display: true,
        grid: {{ display: false }},
        ticks: {{
          maxTicksLimit: 10,
          callback: function(val, idx) {{ return D.cum_labels[idx]; }}
        }}
      }},
      y: {{
        grid: {{ color: '#1f2937' }},
        ticks: {{ callback: v => v + '€' }},
      }}
    }}
  }}
}});

// 1b. Strategy P/L bar
new Chart(document.getElementById('stratChart'), {{
  type: 'bar',
  data: {{
    labels: {strat_names},
    datasets: [{{
      data: {strat_pls},
      backgroundColor: {strat_colors},
      borderRadius: 4,
    }}]
  }},
  options: {{
    indexAxis: 'y',
    plugins: {{
      legend: {{ display: false }},
      tooltip: {{ callbacks: {{ label: ctx => ctx.parsed.x.toFixed(2) + '€' }} }}
    }},
    scales: {{
      x: {{ grid: {{ color: '#1f2937' }}, ticks: {{ callback: v => v + '€' }} }},
      y: {{ grid: {{ display: false }} }}
    }}
  }}
}});

// 2. Monthly P/L
new Chart(document.getElementById('monthlyChart'), {{
  type: 'bar',
  data: {{
    labels: D.monthly_pl.map(m => m.label),
    datasets: [{{
      data: D.monthly_pl.map(m => m.pl),
      backgroundColor: D.monthly_pl.map(m => m.pl >= 0 ? '#10b981' : '#ef4444'),
      borderRadius: 4,
      maxBarThickness: 80,
    }}]
  }},
  options: {{
    plugins: {{
      legend: {{ display: false }},
      tooltip: {{ callbacks: {{ label: ctx => ctx.parsed.y.toFixed(2) + '€ (' + D.monthly_pl[ctx.dataIndex].n + ' bets, WR ' + D.monthly_pl[ctx.dataIndex].wr + '%)' }} }}
    }},
    scales: {{
      x: {{ grid: {{ display: false }} }},
      y: {{ grid: {{ color: '#1f2937' }}, ticks: {{ callback: v => v + '€' }} }}
    }}
  }}
}});

// 3. Minute histogram
new Chart(document.getElementById('minuteChart'), {{
  type: 'bar',
  data: {{
    labels: D.minute_hist.map(m => m.label),
    datasets: [{{
      label: 'N bets',
      data: D.minute_hist.map(m => m.n),
      backgroundColor: D.minute_hist.map(m => m.n === 0 ? '#374151' : m.wr >= 60 ? '#10b981' : m.wr >= 50 ? '#f59e0b' : '#ef4444'),
      borderRadius: 3,
    }}]
  }},
  options: {{
    plugins: {{
      legend: {{ display: false }},
      tooltip: {{ callbacks: {{ label: ctx => ctx.parsed.y + ' bets — WR ' + D.minute_hist[ctx.dataIndex].wr + '%' }} }}
    }},
    scales: {{
      x: {{ grid: {{ display: false }}, title: {{ display: true, text: 'Minuto', color: '#6b7280' }} }},
      y: {{ grid: {{ color: '#1f2937' }} }}
    }}
  }}
}});

// 4. Odds histogram
new Chart(document.getElementById('oddsChart'), {{
  type: 'bar',
  data: {{
    labels: D.odds_hist.map(m => m.label),
    datasets: [{{
      label: 'N bets',
      data: D.odds_hist.map(m => m.n),
      backgroundColor: D.odds_hist.map(m => m.wr >= 60 ? '#10b981' : m.wr >= 50 ? '#f59e0b' : '#ef4444'),
      borderRadius: 3,
    }}]
  }},
  options: {{
    plugins: {{
      legend: {{ display: false }},
      tooltip: {{ callbacks: {{ label: ctx => ctx.parsed.y + ' bets — WR ' + D.odds_hist[ctx.dataIndex].wr + '%' }} }}
    }},
    scales: {{
      x: {{ grid: {{ display: false }}, title: {{ display: true, text: 'Odds', color: '#6b7280' }} }},
      y: {{ grid: {{ color: '#1f2937' }} }}
    }}
  }}
}});

// 5. Day of week (P/L bars)
new Chart(document.getElementById('dowChart'), {{
  type: 'bar',
  data: {{
    labels: D.day_of_week.map(d => d.label),
    datasets: [{{
      data: D.day_of_week.map(d => d.pl),
      backgroundColor: D.day_of_week.map(d => d.pl >= 0 ? '#10b981' : '#ef4444'),
      borderRadius: 3,
    }}]
  }},
  options: {{
    plugins: {{
      legend: {{ display: false }},
      tooltip: {{ callbacks: {{ label: ctx => ctx.parsed.y.toFixed(2) + '€ — ' + D.day_of_week[ctx.dataIndex].n + ' bets — WR ' + D.day_of_week[ctx.dataIndex].wr + '%' }} }}
    }},
    scales: {{ x: {{ grid: {{ display: false }} }}, y: {{ grid: {{ color: '#1f2937' }}, ticks: {{ callback: v => v + '€' }} }} }}
  }}
}});

// 5b. League tier P/L bar
new Chart(document.getElementById('tierPLChart'), {{
  type: 'bar',
  data: {{
    labels: {tier_labels},
    datasets: [{{
      data: {tier_pls},
      backgroundColor: {tier_chart_colors},
      borderRadius: 4,
    }}]
  }},
  options: {{
    indexAxis: 'y',
    plugins: {{
      legend: {{ display: false }},
      tooltip: {{ callbacks: {{ label: ctx => {{
        const i = ctx.dataIndex;
        const tiers = D.league_tiers;
        return ctx.parsed.x.toFixed(2) + '€ — ' + tiers[i].n + ' bets — WR ' + tiers[i].wr + '%';
      }} }} }}
    }},
    scales: {{
      x: {{ grid: {{ color: '#1f2937' }}, ticks: {{ callback: v => v + '€' }} }},
      y: {{ grid: {{ display: false }} }}
    }}
  }}
}});

// 5c. League tier donut
new Chart(document.getElementById('tierDonutChart'), {{
  type: 'doughnut',
  data: {{
    labels: {tier_labels},
    datasets: [{{
      data: {tier_ns},
      backgroundColor: {tier_chart_colors},
      borderColor: '#111827',
      borderWidth: 2,
    }}]
  }},
  options: {{
    plugins: {{
      legend: {{ position: 'bottom', labels: {{ color: '#d1d5db', padding: 16 }} }},
      tooltip: {{ callbacks: {{ label: ctx => {{
        const i = ctx.dataIndex;
        const tiers = D.league_tiers;
        const pct = (tiers[i].n / D.kpis.n_bets * 100).toFixed(1);
        return tiers[i].label + ': ' + tiers[i].n + ' bets (' + pct + '%) — WR ' + tiers[i].wr + '% — P/L ' + tiers[i].pl.toFixed(2) + '€';
      }} }} }}
    }},
  }}
}});

// 6. Hour of day (P/L bars)
new Chart(document.getElementById('hourChart'), {{
  type: 'bar',
  data: {{
    labels: D.hour_of_day.map(h => h.label),
    datasets: [{{
      data: D.hour_of_day.map(h => h.pl),
      backgroundColor: D.hour_of_day.map(h => h.pl >= 0 ? '#10b981' : '#ef4444'),
      borderRadius: 3,
    }}]
  }},
  options: {{
    plugins: {{
      legend: {{ display: false }},
      tooltip: {{ callbacks: {{ label: ctx => ctx.parsed.y.toFixed(2) + '€ — ' + D.hour_of_day[ctx.dataIndex].n + ' bets — WR ' + D.hour_of_day[ctx.dataIndex].wr + '%' }} }}
    }},
    scales: {{ x: {{ grid: {{ display: false }} }}, y: {{ grid: {{ color: '#1f2937' }}, ticks: {{ callback: v => v + '€' }} }} }}
  }}
}});

// ── Table sorting ─────────────────────────────────────────────────────────────
const _tblSort = {{}};
function sortTable(id, col) {{
  const tbl = document.getElementById(id);
  if (!tbl) return;
  const prev = _tblSort[id] || {{col: -1, asc: false}};
  const asc = prev.col === col ? !prev.asc : false;
  _tblSort[id] = {{col, asc}};
  const rows = Array.from(tbl.tBodies[0].rows);
  rows.sort((a, b) => {{
    const av = a.cells[col].dataset.val, bv = b.cells[col].dataset.val;
    const an = parseFloat(av), bn = parseFloat(bv);
    const cmp = isNaN(an) ? av.localeCompare(bv) : an - bn;
    return asc ? cmp : -cmp;
  }});
  rows.forEach(r => tbl.tBodies[0].appendChild(r));
  tbl.querySelectorAll('th .sort-arrow').forEach(s => s.textContent = '');
  const arrow = tbl.tHead.rows[0].cells[col].querySelector('.sort-arrow');
  if (arrow) arrow.textContent = asc ? ' ▲' : ' ▼';
}}
</script>
</body>
</html>'''


# ── Main ────────────────────────────────────────────────────────────────

def main():
    cfg = load_config()

    # Use BT notebook export as canonical source (core + SD, already filtered)
    bets = fetch_bt_bets()
    bets.sort(key=lambda b: b.get("timestamp_utc", ""))

    n_matches = count_matches()
    print(f"  Partidos en data/: {n_matches}")

    team_lookup = load_team_lookup()
    if not team_lookup:
        print("  WARNING: team_lookup.json not found. Run aux/build_team_lookup.py first.")

    # Compute all metrics
    kpis = compute_kpis(bets, n_matches)
    payload = {
        "kpis": kpis,
        "strategy_table": compute_strategy_table(bets),
        "country_table": compute_geo_table(bets, "País"),
        "league_table": compute_geo_table(bets, "Liga"),
        "league_tiers": compute_league_tier_table(bets),
        "tournament_type": compute_tournament_type_table(bets),
        "team_yield": compute_team_yield(bets, team_lookup),
        "day_of_week": compute_day_of_week(bets),
        "hour_of_day": compute_hour_of_day(bets),
        "monthly_pl": compute_monthly_pl(bets),
        "minute_hist": compute_minute_hist(bets),
        "odds_hist": compute_odds_hist(bets),
        "cum_labels": [b.get("timestamp_utc", "")[:10] for b in bets],
        "generated_at": datetime.now().isoformat(),
        "config_summary": {
            "adjustments": cfg.get("adjustments", {}),
            "risk_filter": cfg.get("risk_filter", "all"),
            "preset": cfg.get("active_preset", "none"),
        },
    }

    # ── BackTest page ────────────────────────────────────────────────────
    html = render_html(payload, mode="bt")
    OUTPUT_PATH.write_text(html, encoding="utf-8")
    size_kb = OUTPUT_PATH.stat().st_size / 1024
    print(f"\nBackTest escrito en {OUTPUT_PATH} ({size_kb:.0f} KB)")
    print(f"KPIs BT: {kpis['n_bets']} bets, WR {kpis['wr']}%, ROI {kpis['roi']}%, P/L {kpis['pl']:+.2f}€")

    # ── LIVE page ────────────────────────────────────────────────────────
    live_bets = fetch_live_bets()
    live_bets.sort(key=lambda b: b.get("timestamp_utc", ""))
    live_kpis = compute_kpis(live_bets, n_matches)
    live_payload = {
        "kpis": live_kpis,
        "strategy_table": compute_strategy_table(live_bets),
        "country_table": compute_geo_table(live_bets, "País"),
        "league_table": compute_geo_table(live_bets, "Liga"),
        "league_tiers": compute_league_tier_table(live_bets),
        "tournament_type": compute_tournament_type_table(live_bets),
        "team_yield": compute_team_yield(live_bets, team_lookup),
        "day_of_week": compute_day_of_week(live_bets),
        "hour_of_day": compute_hour_of_day(live_bets),
        "monthly_pl": compute_monthly_pl(live_bets),
        "minute_hist": compute_minute_hist(live_bets),
        "odds_hist": compute_odds_hist(live_bets),
        "cum_labels": [b.get("timestamp_utc", "")[:10] for b in live_bets],
        "generated_at": datetime.now().isoformat(),
        "config_summary": {"adjustments": {}, "risk_filter": "live", "preset": "live"},
    }
    live_html = render_html(live_payload, mode="live")
    LIVE_OUTPUT_PATH.write_text(live_html, encoding="utf-8")
    live_kb = LIVE_OUTPUT_PATH.stat().st_size / 1024
    print(f"LIVE escrito en {LIVE_OUTPUT_PATH} ({live_kb:.0f} KB)")
    if live_bets:
        print(f"KPIs LIVE: {live_kpis['n_bets']} bets, WR {live_kpis['wr']}%, ROI {live_kpis['roi']}%, P/L {live_kpis['pl']:+.2f}€")


if __name__ == "__main__":
    main()
