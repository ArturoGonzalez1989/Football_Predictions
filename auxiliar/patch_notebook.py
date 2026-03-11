import sys, json
sys.stdout.reconfigure(encoding='utf-8')

nb_path = r'C:\Users\agonz\OneDrive\Documents\Proyectos\Furbo\analisis\reconcile_bt_live.ipynb'
nb = json.load(open(nb_path, encoding='utf-8'))

# ── Cell 2 ──
cell2 = """\
# ── Celda 2: Backtest — leer portfolio_bets_*.xlsx (output final del notebook) ──
import openpyxl as _xl2
import pandas as pd
from pathlib import Path as _P2
from collections import Counter as _Counter2

ANALISIS_DIR = ROOT / 'analisis'

_xl_candidates = sorted(ANALISIS_DIR.glob('portfolio_bets_*.xlsx'), reverse=True)
if not _xl_candidates:
    raise FileNotFoundError("No se encontro portfolio_bets_*.xlsx. Ejecuta strategies_designer.ipynb primero.")

BT_EXCEL_PATH = _xl_candidates[0]
print(f'Leyendo BT Excel: {BT_EXCEL_PATH.name}')

_wb = _xl2.load_workbook(BT_EXCEL_PATH, read_only=True, data_only=True)
_ws = _wb.active
_headers = [c.value for c in next(_ws.iter_rows(min_row=1, max_row=1))]
_bt_rows = [
    {_headers[i]: (c.value if c.value is not None else '') for i, c in enumerate(row)}
    for row in _ws.iter_rows(min_row=2)
]
print(f'BT bets totales: {len(_bt_rows)}')

def _norm_strat(s):
    s = str(s or '')
    for sfx in ['_v1','_v2','_v3','_v4','_v15','_v2r','_base']:
        if s.endswith(sfx):
            return s[:-len(sfx)]
    return s

_bt_records = []
for r in _bt_rows:
    _bt_records.append({
        'match_id':        str(r.get('match_id', '')),
        'strategy_family': _norm_strat(r.get('strategy', '')),
        'strategy_raw':    str(r.get('strategy', '')),
        'minuto_bt':       r.get('minuto'),
        'odds_bt':         r.get('effective_odds') or r.get('back_odds'),
        'won_bt':          r.get('won') in (True, 'True', 'true', 1, '1'),
        'pl_bt':           float(r.get('pl_eur') or 0),
        'timestamp_utc':   str(r.get('timestamp_utc', '')),
    })

df_bt_raw = pd.DataFrame(_bt_records)
print(f'df_bt_raw shape: {df_bt_raw.shape}')
print('Estrategias en BT:')
for k, v in sorted(_Counter2(df_bt_raw['strategy_raw']).items(), key=lambda x: -x[1]):
    print(f'  {k}: {v}')
print(f'Rango fechas BT: {df_bt_raw["timestamp_utc"].min()} -> {df_bt_raw["timestamp_utc"].max()}')
"""

# ── Cell 10 ──
cell10 = """\
# ── Celda 9: Paper Bets vs BT Excel ──
# Fuente BT   : portfolio_bets_*.xlsx (filtros finales aplicados por notebook)
# Fuente PAPER: placed_bets.csv (apuestas paper trading reales)
# Filtra BT desde la fecha de inicio del paper trading
import csv as _csv9
import pandas as pd
from pathlib import Path as _P9
from collections import defaultdict as _dd9

PLACED_BETS_CSV = ROOT / 'betfair_scraper' / 'placed_bets.csv'

paper_bets = []
if PLACED_BETS_CSV.exists():
    with open(PLACED_BETS_CSV, 'r', encoding='utf-8') as f:
        for row in _csv9.DictReader(f):
            paper_bets.append(row)

if not paper_bets:
    print('No hay apuestas en placed_bets.csv todavia.')
else:
    _paper_dates = [r['timestamp_utc'] for r in paper_bets if r.get('timestamp_utc')]
    paper_start = min(_paper_dates)
    print(f'Paper trading arranco: {paper_start}')

    df_bt_period = df_bt_raw[df_bt_raw['timestamp_utc'] >= paper_start[:10]].copy()
    print(f'BT bets desde {paper_start[:10]}: {len(df_bt_period)}')

    def _paper_norm(s):
        s = str(s or '')
        for sfx in ['_v1','_v2','_v3','_v4','_v15','_v2r','_base']:
            if s.endswith(sfx):
                return s[:-len(sfx)]
        return s

    paper_settled = [r for r in paper_bets if r.get('status') != 'pending']
    print(f'Paper bets (excl pending): {len(paper_settled)}')

    df_paper = pd.DataFrame([{
        'match_id':        r.get('match_id',''),
        'match_name':      r.get('match_name',''),
        'strategy_family': _paper_norm(r.get('strategy','')),
        'strategy_raw':    r.get('strategy',''),
        'minute_paper':    float(r['minute']) if r.get('minute') else None,
        'odds_paper':      float(r['back_odds']) if r.get('back_odds') else None,
        'status':          r.get('status',''),
        'result':          r.get('result',''),
        'pl':              float(r['pl']) if r.get('pl') else None,
        'score':           r.get('score',''),
        'timestamp':       r.get('timestamp_utc',''),
    } for r in paper_settled])

    bt_keys   = _dd9(list)
    for _, r in df_bt_period.iterrows():
        bt_keys[(r['match_id'], r['strategy_family'])].append(r)

    paper_keys = _dd9(list)
    for _, r in df_paper.iterrows():
        paper_keys[(r['match_id'], r['strategy_family'])].append(r)

    in_both    = [(k, bt_keys[k], paper_keys[k]) for k in bt_keys if k in paper_keys]
    bt_only    = [(k, bt_keys[k]) for k in bt_keys if k not in paper_keys]
    paper_only = [(k, paper_keys[k]) for k in paper_keys if k not in bt_keys]

    SEP = '=' * 65
    print(f'\\n{SEP}')
    print('RESUMEN')
    print(SEP)
    print(f'  En ambos  (BT+PAPER): {len(in_both):>3}')
    print(f'  Solo en BT           : {len(bt_only):>3}  <- paper deberia haberlos apostado')
    print(f'  Solo en PAPER        : {len(paper_only):>3}  <- paper aposto algo fuera de BT')

    if bt_only:
        print(f'\\n{SEP}')
        print(f'SOLO EN BT ({len(bt_only)}) -- no apostados en paper:')
        print(SEP)
        for (mid, strat), rows in sorted(bt_only, key=lambda x: x[0][1]):
            r = rows[0]
            won_str = 'W' if r['won_bt'] else 'L'
            print(f'  {mid}')
            print(f'    {str(r["strategy_raw"]):<35s} min={str(r["minuto_bt"]):>4}  {won_str}  pl={r["pl_bt"]:+.3f}')

    if paper_only:
        print(f'\\n{SEP}')
        print(f'SOLO EN PAPER ({len(paper_only)}) -- no validados por BT:')
        print(SEP)
        for (mid, strat), rows in sorted(paper_only, key=lambda x: x[0][1]):
            r = rows[0]
            pl_str = f'{r["pl"]:+.2f}' if pd.notna(r['pl']) else 'pending'
            print(f'  {str(r["match_name"]):<45s} ({r["score"]})')
            print(f'    {str(r["strategy_raw"]):<35s} min={str(r["minute_paper"] or ""):>4}  odds={str(r["odds_paper"] or ""):>5}  -> {r["result"]:<4}  pl={pl_str}')

    print(f'\\n{SEP}')
    print('P/L COMPARATIVO (stake=1 EUR)')
    print(SEP)
    bt_pl  = df_bt_period['pl_bt'].sum()
    pap_pl = df_paper['pl'].sum()
    bt_wr  = df_bt_period['won_bt'].mean() * 100
    pap_wr = (df_paper['result'] == 'won').mean() * 100
    print(f'  BT    ({len(df_bt_period):>3} bets): WR={bt_wr:.1f}%  P/L={bt_pl:+.2f}')
    print(f'  PAPER ({len(df_paper):>3} bets): WR={pap_wr:.1f}%  P/L={pap_pl:+.2f}')
"""

nb['cells'][2]['source'] = [cell2]
nb['cells'][10]['source'] = [cell10]

open(nb_path, 'w', encoding='utf-8').write(json.dumps(nb, ensure_ascii=False, indent=1))
print("Notebook guardado OK.")
