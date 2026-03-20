"""
bt_optimizer.py — Backtest optimizer for Furbo betting strategies.

Replaces the strategies_designer.ipynb notebook with a fully reproducible CLI script.

Phases:
  0  Pre-load all historical match data into memory (once, ~5 s)
  1  Grid search per strategy → best (params) using _analyze_strategy_simple directly
  2  Quality gate: discard strategies that don't pass N / ROI / IC95 thresholds
  3  Portfolio preset generation: run optimizer_cli.run() for 4 criteria
  4  Select best preset and apply to cartera_config.json
  5  Export CSV + XLSX with all bets from the final portfolio
  6  Monte Carlo risk analysis: strategy fragility, DD distribution, profit distribution

Usage:
  python scripts/bt_optimizer.py                      # run all phases
  python scripts/bt_optimizer.py --phase individual   # only phase 1+2
  python scripts/bt_optimizer.py --phase presets      # only phase 3
  python scripts/bt_optimizer.py --phase apply --criterion max_pl
  python scripts/bt_optimizer.py --phase export
  python scripts/bt_optimizer.py --dry-run            # never writes cartera_config.json
  python scripts/bt_optimizer.py --workers 8
"""

import sys
import io
# Force UTF-8 output on Windows (avoids UnicodeEncodeError for → ✓ etc.)
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
else:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
import json
import math
import shutil
import argparse
import itertools
import time
from pathlib import Path
from datetime import datetime

# ── Path setup ────────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent
BACKEND = ROOT / "betfair_scraper" / "dashboard" / "backend"
sys.path.insert(0, str(BACKEND))
sys.path.insert(0, str(ROOT / "betfair_scraper"))

from utils import csv_reader
from utils.csv_reader import (
    _STRATEGY_REGISTRY,
    _STRATEGY_REGISTRY_KEYS,
    _analyze_strategy_simple,
    _cfg_add_snake_keys,
    _normalize_mercado,
    _TO_CANONICAL,
)
from utils.csv_loader import _get_all_finished_matches, _read_csv_rows, _final_result_row
from api import optimizer_cli
import monte_carlo

# ── Paths ─────────────────────────────────────────────────────────────────────
CARTERA_CFG   = ROOT / "betfair_scraper" / "cartera_config.json"
CARTERA_BAK_DIR = ROOT / "backup" / "cartera_config"
PRESETS_DIR   = ROOT / "betfair_scraper" / "data" / "presets"
EXPORTS_DIR   = ROOT / "analisis"
RESULTS_FILE  = ROOT / "auxiliar" / "bt_optimizer_results.json"

PRESETS_DIR.mkdir(parents=True, exist_ok=True)
CARTERA_BAK_DIR.mkdir(parents=True, exist_ok=True)

# ── Estrategias permanentemente excluidas del optimizer ───────────────────────
# Riesgo asimétrico LAY: pérdidas de hasta (odds-1)x stake vs ganancia máx 1x stake
_PERMANENTLY_DISABLED: set[str] = {
    "lay_over45_v3",
    "lay_over45_blowout",
    "lay_cs11",
}

# ── Quality gates ─────────────────────────────────────────────────────────────
G_MIN_ROI        = 10.0   # minimum ROI %
IC95_MIN_LOW     = 40.0   # minimum Wilson CI lower bound
def _min_pl_per_bet(n_fin: int) -> float:
    """Dynamic PL/bet threshold: more data = more demanding. Capped at 0.30."""
    return min(0.30, 0.10 + n_fin / 10000)

# ── Selector ──────────────────────────────────────────────────────────────────
CRITERIA = ["max_roi", "max_pl", "max_wr", "min_dd"]
DEFAULT_SELECTOR = "robust"          # ci_low × wr × sqrt(N) — rewards quality + sample size
MIN_PRESET_N     = 200  # real stats from analyze_cartera-equivalent; N<200 signals something odd

# ── Min-duration defaults (same as cartera_config.json) ──────────────────────
DEFAULT_MIN_DUR = {
    "back_draw_00":        2,
    "xg_underperformance": 3,
    "odds_drift":          2,
    "goal_clustering":     4,
    "pressure_cooker":     4,
}

# ─────────────────────────────────────────────────────────────────────────────
# SEARCH SPACES
# Keys are snake_case — must match what trigger functions read via cfg.get().
# Multi-name params (m_min vs minute_min) are normalized to canonical form
# by _TO_CANONICAL when writing to cartera_config.json.
# ─────────────────────────────────────────────────────────────────────────────
SEARCH_SPACES: dict[str, dict[str, list]] = {

    # ── Original 7 strategies ─────────────────────────────────────────────────

    "back_draw_00": {
        "xg_max":    [0.4, 0.5, 0.6, 0.7, 1.0],
        "poss_max":  [15, 20, 25, 30, 100],
        "shots_max": [6, 8, 10, 12, 20],
        "minute_min":[25, 30, 35],
        "minute_max":[75, 80, 85, 90],
    },
    "xg_underperformance": {           "xg_excess_min": [0.3, 0.4, 0.5, 0.6, 0.7],
        "sot_min":       [0, 1, 2, 3],
        "minute_min":    [0, 10, 15, 20],
        "minute_max":    [60, 65, 70, 75, 80, 90],
        "odds_min":      [0, 1.20, 1.30, 1.40],
    },
    "odds_drift": {           "drift_min_pct": [15, 20, 30],
        "max_odds":      [5.0, 10.0, 15.0, 20.0, 999],
        "goal_diff_min": [0, 1],
        "min_minute":    [0, 30],
        "max_minute":    [85, 90],
        "odds_min":      [0, 1.20, 1.40],
    },
    "goal_clustering": {
        "sot_min":    [0, 2, 3, 4],
        "min_minute": [0, 15, 20],
        "max_minute": [55, 60, 65, 70],
        "xg_rem_min": [0.0, 0.3, 0.6],
        "odds_min":   [0, 1.20, 1.30],
    },
    "pressure_cooker": {
        "min_minute":    [55, 60, 65, 68],
        "max_minute":    [70, 73, 75, 78],
        "score_confirm": [1, 2, 3],
        "odds_min":      [0, 1.20, 1.30],
    },
    "momentum_xg": {           "min_m": [0, 5, 10, 15, 20],
        "max_m": [70, 75, 80, 85, 90],
    },

    # ── 19 additional strategies ───────────────────────────────────────────────
    "over25_2goal": {
        "m_min":         [45, 50, 55, 60],
        "m_max":         [75, 78, 81, 84],
        "goal_diff_min": [1, 2],
        "sot_total_min": [0, 2, 3, 5],
        "odds_min":      [1.2, 1.4, 1.6],
        "odds_max":      [6.0, 8.0, 10.0],
    },
    "under35_late": {
        "m_min":      [60, 65, 70],
        "m_max":      [75, 78, 81, 85],
        "xg_max":     [2.5, 3.0, 3.5],
        "goals_min":  [2, 3],
        "goals_max":  [3, 4],
        "odds_min":   [0, 1.20, 1.30, 1.40],
        "odds_max":   [6.0, 8.0, 10.0],
    },
    "longshot": {
        "m_min":     [60, 65, 70],
        "m_max":     [80, 83, 86, 88],
        "odds_min":  [1.2, 1.3, 1.5],
        "odds_max":  [5.0, 7.0, 10.0],
    },
    "cs_close": {
        "m_min":    [60, 65, 67, 70],
        "m_max":    [78, 80, 83, 85],
        "odds_min": [0, 3.0, 4.0, 5.0],
        "odds_max": [8.0, 10.0, 12.0, 15.0, 999],
    },
    "cs_one_goal": {
        "m_min":    [60, 65, 68, 70],
        "m_max":    [82, 85, 88, 90],
        "odds_min": [0, 2.5, 3.0, 3.5],
        "odds_max": [8.0, 10.0, 12.0, 15.0, 999],
    },
    "ud_leading": {
        "m_min":            [45, 50, 53, 58],
        "m_max":            [78, 80, 83, 85],
        "ud_min_pre_odds":  [1.3, 1.5, 1.8, 2.0, 2.5],
        "max_lead":         [1, 2, 3],
        "odds_min":         [0, 1.20, 1.40, 1.60],
    },
    "home_fav_leading": {
        "m_min":     [50, 55, 60],
        "m_max":     [75, 78, 80, 83],
        "max_lead":  [1, 2, 3],
        "fav_max":   [2.0, 2.5, 3.0],
        "odds_min":  [0, 1.20, 1.30, 1.40],
    },
    "cs_20": {
        "m_min":    [65, 70, 72, 75],
        "m_max":    [83, 85, 88, 90],
        "odds_max": [8.0, 10.0, 15.0],
        "odds_min": [0, 2.0, 3.0, 4.0],
    },
    "cs_big_lead": {
        "m_min":    [65, 68, 70, 72],
        "m_max":    [80, 82, 85],
        "odds_max": [6.0, 8.0, 10.0],
        "odds_min": [0, 2.0, 3.0, 4.0],
    },
    "lay_over45_v3": {
        "m_min":      [45, 50, 55, 60],
        "m_max":      [70, 73, 75, 78],
        "goals_max":  [1, 2, 3],
        "odds_max":   [10.0, 15.0, 20.0],
    },
    "draw_xg_conv": {
        "m_min":       [55, 60, 63, 65],
        "m_max":       [78, 80, 83, 85],
        "xg_diff_max": [0.5, 0.8, 1.0, 1.3],
        "odds_min":    [0, 1.50, 1.80, 2.10],
    },
    "poss_extreme": {
        "m_min":    [25, 30, 35],
        "m_max":    [45, 50, 53, 55],
        "poss_min": [50, 55, 58, 60],
    },
    "cs_00": {
        "m_min":    [25, 28, 30],
        "m_max":    [30, 33, 35],
        "xg_max":   [1.0, 1.5, 2.0],
        "odds_min": [4.0, 5.0, 6.0],
        "odds_max": [10.0, 12.0, 15.0],
    },
    "over25_2goals": {
        "m_min":    [40, 45, 48, 52],
        "m_max":    [58, 60, 63, 65],
        "odds_max": [3.0, 4.0, 5.0, 6.0],
        "odds_min": [0, 1.20, 1.30, 1.40],
    },
    "draw_11": {
        "m_min":    [62, 65, 68, 70],
        "m_max":    [82, 85, 88, 90],
        "odds_min": [1.0, 1.2, 1.4],
    },
    "under35_3goals": {
        "m_min":    [60, 65, 70],
        "m_max":    [80, 83, 85, 90],
        "xg_max":   [2.0, 2.5, 3.0],
        "odds_min": [0, 1.20, 1.30, 1.40],
    },
    "away_fav_leading": {
        "m_min":    [60, 65, 68, 70],
        "m_max":    [80, 83, 85, 88],
        "max_lead": [1, 2, 3],
        "fav_max":  [2.0, 2.5, 3.0],
        "odds_min": [0, 1.20, 1.40, 1.60],
        "odds_max": [5.0, 8.0, 10.0],
    },
    "under45_3goals": {
        "m_min":    [50, 55, 58, 60],
        "m_max":    [75, 78, 80, 83],
        "xg_max":   [2.0, 2.5, 3.0],
        "odds_min": [0, 1.20, 1.30, 1.40],
    },
    "cs_11": {
        "m_min":    [62, 65, 68, 70],
        "m_max":    [82, 85, 88, 90],
        "odds_min": [0, 3.0, 4.0, 5.0],
        "odds_max": [8.0, 12.0, 15.0, 999],
    },
    "draw_equalizer": {
        "m_min":          [60, 63, 65, 68],
        "m_max":          [85, 88, 90],
        "fav_pre_max":    [2.0, 2.5, 3.0],
        "min_goals_each": [1, 2],
        "odds_max":       [6.0, 8.0, 10.0],
        "odds_min":       [0, 1.50, 2.00, 2.50],
    },
    "draw_22": {
        "m_min":    [65, 68, 70, 72],
        "m_max":    [85, 88, 90],
        "odds_max": [6.0, 8.0, 10.0],
        "odds_min": [0, 1.50, 2.00, 2.50],
    },
    "lay_over45_blowout": {
        "m_min":        [55, 58, 60, 63],
        "m_max":        [72, 75, 78],
        "post_window":  [5, 8, 10, 12],
        "sot_max":      [0, 1, 2],
        "odds_max":     [10.0, 15.0, 20.0],
        "include_31":   [0, 1],
    },

    # ── R19 brute-force discoveries ────────────────────────────────────────────
    "over35_early_goals": {
        "m_min":     [40, 45, 50],
        "m_max":     [60, 65, 70],
        "odds_min":  [1.5, 1.8, 2.0],
        "odds_max":  [6.0, 8.0, 10.0],
    },
    "lay_draw_away_leading": {  # LAY — win=+stake*0.95, loss=-stake*(odds-1)
        "m_min":    [50, 55, 60],
        "m_max":    [75, 78, 80],
        "xg_max":   [1.2, 1.5, 1.8, 2.0],
        "odds_min": [2.0, 2.5],
        "odds_max": [8.0, 10.0],
    },
    "lay_cs11": {               # LAY — win=+stake*0.95, loss=-stake*(odds-1)
        "m_min":     [55, 58, 60, 63],
        "m_max":     [80, 83, 85, 88],
        "odds_min":  [1.5, 2.0],
        "odds_max":  [20.0, 50.0],
    },

    # tarde_asia: no tunable params (liga-based detection), evaluated once with defaults
    "tarde_asia": {},
}

# All strategy keys — one registry entry per strategy, no versioning, no categories.
SINGLE_STRATEGIES = [
    "back_draw_00", "xg_underperformance", "odds_drift", "momentum_xg",
    "goal_clustering", "pressure_cooker",
    "over25_2goal", "under35_late", "longshot",
    "cs_close", "cs_one_goal", "ud_leading", "home_fav_leading",
    "cs_20", "cs_big_lead", "lay_over45_v3", "draw_xg_conv",
    "poss_extreme", "cs_00", "over25_2goals", "draw_11",
    "under35_3goals", "away_fav_leading", "under45_3goals", "cs_11",
    "draw_equalizer", "draw_22",
    "lay_over45_blowout",
    "over35_early_goals", "lay_draw_away_leading", "lay_cs11",
    "tarde_asia",
]


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _wilson_ci(wins: int, n: int, z: float = 1.96):
    if n == 0:
        return (0.0, 0.0)
    p = wins / n
    center = (p + z * z / (2 * n)) / (1 + z * z / n)
    margin  = z * math.sqrt(p * (1-p) / n + z*z / (4*n*n)) / (1 + z*z/n)
    return (round(max(0.0, center - margin) * 100, 1),
            round(min(1.0, center + margin) * 100, 1))


def _n_matches() -> int:
    return len(_get_all_finished_matches())


def _min_n(n_fin: int) -> int:
    return max(15, n_fin // 25)


def _eval_bets(bets: list, n_fin: int) -> dict | None:
    """Apply quality gates; return metrics dict or None if gates fail."""
    n = len(bets)
    if n < _min_n(n_fin):
        return None
    wins = sum(1 for b in bets if b["won"])
    pl   = sum(b["pl"] for b in bets)
    roi  = pl / n * 100
    if roi < G_MIN_ROI:
        return None
    if pl / n < _min_pl_per_bet(n_fin):
        return None
    ci_l, ci_h = _wilson_ci(wins, n)
    if ci_l < IC95_MIN_LOW:
        return None
    max_dd = _max_drawdown([b["pl"] for b in bets])
    return {
        "n":      n,
        "wins":   wins,
        "wr":     round(wins / n * 100, 1),
        "pl":     round(pl, 2),
        "roi":    round(roi, 1),
        "ci_low": ci_l,
        "ci_high":ci_h,
        "max_dd": max_dd,
        "score":  round(ci_l * roi / 100, 4),
    }


def _max_drawdown(pl_list: list) -> float:
    peak = cur = 0.0
    max_dd = 0.0
    for pl in pl_list:
        cur += pl
        if cur > peak:
            peak = cur
        dd = peak - cur
        if dd > max_dd:
            max_dd = dd
    return round(max_dd, 2)


def _eval_on_matches_subset(key: str, entry: tuple, cfg: dict, min_dur: int,
                            matches: list) -> list:
    """Run a strategy on a specific subset of matches. Used by Phase 2.5 crossval.

    Uses _final_result_row() for FT score extraction — same logic as
    _analyze_strategy_simple() — to avoid divergence on matches with
    trailing rows or missing 'finalizado' estado.
    """
    _, _, trigger_fn, _, extract_fn, win_fn = entry
    bets = []
    for match_data in matches:
        rows = match_data.get("rows") or _read_csv_rows(match_data["csv_path"])
        if not rows:
            continue
        match_id = match_data["match_id"]
        last = _final_result_row(rows)
        if last is None:
            continue
        try:
            ft_gl = int(float(last["goles_local"]))
            ft_gv = int(float(last["goles_visitante"]))
        except (ValueError, TypeError):
            continue
        effective_cfg = {**cfg, "match_id": match_id,
                         "match_name": match_data.get("name", ""),
                         "match_url":  match_data.get("url", "")}
        first_seen = None
        trig_data  = None
        for curr_idx in range(len(rows)):
            trig = trigger_fn(rows, curr_idx, effective_cfg)
            if trig:
                if first_seen is None:
                    first_seen = curr_idx
                    trig_data  = trig
                if curr_idx >= first_seen + min_dur - 1:
                    extracted = extract_fn(trig_data)
                    if extracted is None:
                        break
                    odds, rec, _ = extracted
                    won   = win_fn(trig_data, ft_gl, ft_gv)
                    is_lay = rec.upper().startswith("LAY")
                    pl = round(0.95 if won else -(odds - 1), 2) if is_lay \
                         else round((odds - 1) * 0.95 if won else -1.0, 2)
                    bets.append({"won": won, "pl": pl})
                    break
    return bets


def _crossval_raw_metrics(bets: list) -> dict:
    """Raw metrics without quality gates — used for per-fold reporting."""
    n = len(bets)
    if n == 0:
        return {"n": 0, "wr": 0.0, "roi": 0.0}
    wins = sum(1 for b in bets if b["won"])
    pl   = sum(b["pl"] for b in bets)
    return {"n": n, "wr": round(wins / n * 100, 1), "roi": round(pl / n * 100, 1)}


def _registry_entry(key: str):
    for e in _STRATEGY_REGISTRY:
        if e[0] == key:
            return e
    return None


def _get_min_dur(key: str, cfg_md: dict) -> int:
    return cfg_md.get(key, 1)


def _load_config() -> dict:
    with open(CARTERA_CFG, encoding="utf-8") as f:
        return json.load(f)


def _log(msg: str):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


# ─────────────────────────────────────────────────────────────────────────────
# PHASE 0 — Pre-load match data
# ─────────────────────────────────────────────────────────────────────────────

def phase0_load() -> tuple[list, int]:
    """Load all finished matches into memory cache. Returns (matches, n_matches)."""
    _log("Phase 0 — cargando datos históricos…")
    t0 = time.time()
    matches = _get_all_finished_matches()
    n = len(matches)
    _log(f"  {n} partidos cargados en {time.time()-t0:.1f}s")
    return matches, n


# ─────────────────────────────────────────────────────────────────────────────
# PHASE 1.5 — Odds-minimum calibration (analytical edge threshold)
# ─────────────────────────────────────────────────────────────────────────────

# Buckets: (lower_inclusive, upper_exclusive, midpoint_for_implied_wr)
# Midpoint used to compute market-implied win probability (1/midpoint).
_ODDS_BUCKETS = [
    (1.00, 1.30, 1.15),
    (1.30, 1.50, 1.40),
    (1.50, 1.65, 1.575),
    (1.65, 1.80, 1.725),
    (1.80, 2.10, 1.95),
    (2.10, 3.00, 2.55),
    (3.00, float("inf"), 4.50),
]
_MIN_BUCKET_N = 5  # minimum bets per bucket to trust its win-rate estimate


def _calibrate_odds_min(key: str, best_params: dict, min_dur: int) -> float | None:
    """Determine the minimum entry odds for a strategy from first principles.

    Runs the strategy with odds_min=0 (no filter) using the best non-odds params
    found by the grid search, then computes actual WR per odds bucket and compares
    it to the market-implied win probability (1 / bucket_midpoint_odds).

    Returns the lower bound of the lowest bucket where actual WR > implied WR
    (i.e. where the strategy has a positive edge), or 0.0 if edge exists even
    at very low odds.  Returns None if calibration does not apply (LAY strategy,
    no data, etc.).

    Why this complements the grid search:
      The grid maximises a composite score (ci_low × roi) over a coarse discrete
      search space.  This function instead asks the fundamental question: "at what
      odds level does our historical win-rate exceed the market's implied
      probability?"  The answer is strategy-specific and data-driven rather than
      dependent on the grid's granularity.
    """
    # LAY strategies have asymmetric P/L — WR vs implied-WR comparison is invalid.
    if "lay" in key.lower() or key in _PERMANENTLY_DISABLED:
        return None

    entry = _registry_entry(key)
    if entry is None:
        return None
    _, _, trigger_fn, _, extract_fn, win_fn = entry

    # Run with no odds floor (odds_min=0) using all other best params.
    params_no_odds = {k: v for k, v in best_params.items()
                      if k not in ("odds_min", "min_odds")}
    cfg = _cfg_add_snake_keys({**params_no_odds, "odds_min": 0, "enabled": True})
    bets = _analyze_strategy_simple(key, trigger_fn, extract_fn, win_fn, cfg, min_dur)
    if not bets:
        return None

    # Compute WR and edge per odds bucket (scan from lowest to highest odds).
    _first_bucket_lo = _ODDS_BUCKETS[0][0]
    calibrated_min = 0.0
    found_any_bucket = False
    for lo, hi, mid in _ODDS_BUCKETS:
        bucket = [b for b in bets if lo <= (b.get("back_odds") or 0) < hi]
        n = len(bucket)
        if n < _MIN_BUCKET_N:
            # Too few bets to be statistically meaningful — skip this bucket.
            continue
        found_any_bucket = True
        wr = sum(1 for b in bucket if b["won"]) / n
        implied_wr = 1.0 / mid
        edge = wr - implied_wr
        if edge > 0:
            # First bucket with positive edge found.
            # If it's the very first bucket (lo == 1.0), no minimum is needed.
            calibrated_min = 0.0 if lo <= _first_bucket_lo else lo
            break
        else:
            # No edge at these odds — minimum must be at least the top of this bucket.
            calibrated_min = hi if hi != float("inf") else calibrated_min

    return round(calibrated_min, 2)


# ─────────────────────────────────────────────────────────────────────────────
# PHASE 1 — Individual grid search
# ─────────────────────────────────────────────────────────────────────────────

def _run_single_strategy(key: str, space: dict, min_dur: int, n_fin: int) -> dict | None:
    """Grid search for a single (non-versioned) strategy. Returns best result or None."""
    entry = _registry_entry(key)
    if entry is None:
        return None
    _, _, trigger_fn, _, extract_fn, win_fn = entry

    param_names  = list(space.keys())
    param_values = list(space.values())
    best = None

    for combo_vals in itertools.product(*param_values):
        raw_cfg = dict(zip(param_names, combo_vals))
        cfg     = _cfg_add_snake_keys({**raw_cfg, "enabled": True})
        bets    = _analyze_strategy_simple(key, trigger_fn, extract_fn, win_fn, cfg, min_dur)
        metrics = _eval_bets(bets, n_fin)
        if metrics and (best is None or metrics["score"] > best["score"]):
            best = {**metrics, "params": raw_cfg, "key": key}

    return best


def phase1_individual(n_fin: int, workers: int = 4,
                       only: list | None = None) -> dict:
    """
    Run grid search for all strategies. Returns dict keyed by strategy key:
      { "goal_clustering": { n, wins, wr, pl, roi, ci_low, ci_high, max_dd, score,
                              params: {...}, key: "goal_clustering" },
        "back_draw_00":    { ..., key: "back_draw_00", params: {...} },
        ... }
    Strategies that don't pass quality gates are absent from the result.
    """
    cfg   = _load_config()
    md    = cfg.get("min_duration", {})
    results: dict[str, dict] = {}

    _log(f"Phase 1 — grid search ({n_fin} partidos)…")
    tasks = []

    for key in SINGLE_STRATEGIES:
        if only and key not in only:
            continue
        if key in _PERMANENTLY_DISABLED:
            continue
        space = SEARCH_SPACES.get(key)
        if space is None:
            _log(f"  WARN: {key} — sin search space definido, skip")
            continue
        tasks.append((key, space, _get_min_dur(key, md)))

    n_total = len(tasks)
    t0      = time.time()

    # Sequential execution — trigger lambdas in the registry are not picklable.
    # Match data is in module cache so each call is fast (~0.3-0.5 s).
    for key, space, min_dur in tasks:
        best = _run_single_strategy(key, space, min_dur, n_fin)
        if best:
            # Phase 1.5 — refine odds_min analytically (WR vs implied WR per bucket).
            cal_min = _calibrate_odds_min(key, best["params"], min_dur)
            if cal_min is not None:
                grid_min = best["params"].get("odds_min", 0)
                if cal_min != grid_min:
                    # Validate the calibrated minimum still passes quality gates.
                    entry = _registry_entry(key)
                    if entry is not None:
                        _, _, trigger_fn, _, extract_fn, win_fn = entry
                        cal_cfg = _cfg_add_snake_keys(
                            {**best["params"], "odds_min": cal_min, "enabled": True}
                        )
                        cal_bets = _analyze_strategy_simple(
                            key, trigger_fn, extract_fn, win_fn, cal_cfg, min_dur
                        )
                        cal_metrics = _eval_bets(cal_bets, n_fin)
                        if cal_metrics:
                            _log(f"    → oddsMin calibrado: {grid_min} → {cal_min} "
                                 f"(N={cal_metrics['n']} ROI={cal_metrics['roi']}% "
                                 f"WR={cal_metrics['wr']}%)")
                            best["params"]["odds_min"] = cal_min
                            best.update(cal_metrics)
                        else:
                            _log(f"    → oddsMin calibrado {cal_min} no pasa quality gates "
                                 f"— se mantiene grid: {grid_min}")

            results[key] = best
            _log(f"  ✓ {key}: N={best['n']} ROI={best['roi']}% "
                 f"IC=[{best['ci_low']}-{best['ci_high']}] "
                 f"params={best['params']}")
        else:
            _log(f"  ✗ {key}: no pasó quality gates")

    elapsed = time.time() - t0
    _log(f"Phase 1 completada en {elapsed:.1f}s — "
         f"{len(results)}/{n_total} estrategias aprobadas")
    return results


# ─────────────────────────────────────────────────────────────────────────────
# PHASE 2 — Build optimal config from individual results
# ─────────────────────────────────────────────────────────────────────────────

def phase2_build_config(individual_results: dict) -> dict:
    """
    Merge best individual params into cartera_config.json format.
    Returns the new strategies dict (does NOT write to disk yet).
    """
    _log("Phase 2 — construyendo config óptimo por estrategia…")
    base_cfg = _load_config()
    new_strategies = dict(base_cfg.get("strategies", {}))

    approved   = []
    discarded  = []

    for family, result in individual_results.items():
        params = result["params"]
        canonical_params = _normalize_params(params)
        new_strategies[family] = {
            "enabled": True,
            **canonical_params,
            "_stats": {
                "wr": round(result["wr"], 1),
                "roi": round(result["roi"], 1),
                "n": result["n"],
                "ci_low": round(result["ci_low"], 1),
            },
        }
        approved.append(family)

    # Disable strategies that didn't pass quality gates
    for fam in SINGLE_STRATEGIES:
        if fam not in individual_results:
            if fam in new_strategies:
                new_strategies[fam] = {**new_strategies[fam], "enabled": False}
                discarded.append(fam)

    _log(f"  Aprobadas ({len(approved)}): {', '.join(approved)}")
    if discarded:
        _log(f"  Descartadas ({len(discarded)}): {', '.join(discarded)}")
    return new_strategies


def _normalize_params(params: dict) -> dict:
    """Normalize param key variants to canonical snake_case for cartera_config.json.

    Uses _TO_CANONICAL from csv_reader to map aliases (m_min, min_m, min_minute)
    to their canonical form (minute_min). Keys without aliases pass through as-is.
    """
    return {_TO_CANONICAL.get(k, k): v for k, v in params.items()}


# ─────────────────────────────────────────────────────────────────────────────
# PHASE 2.5 — Cross-validation robustness filter
# ─────────────────────────────────────────────────────────────────────────────

def phase2_5_crossval(individual_results: dict, new_strategies: dict,
                      all_matches: list, k: int = 5, seed: int = 42
                      ) -> tuple[dict, dict]:
    """
    Validate that strategies which passed quality gates (Phase 1+2) are robust
    across different match subsets. Fragile strategies are removed from
    individual_results and disabled in new_strategies before Phase 3.

    Robustness criteria:
      - mean fold ROI >= G_MIN_ROI (10%)
      - >= 60% of folds have positive ROI
      - temporal test ROI >= 0% (last 30% of matches by date)

    Returns (updated_individual_results, updated_new_strategies).
    """
    import math as _math
    _log("Phase 2.5 — cross-validation de robustez…")

    if not individual_results:
        _log("  Nada que validar — saltando Phase 2.5")
        return individual_results, new_strategies

    # Sort by date for temporal split
    def _match_date(m):
        rows = m.get("rows")
        if not rows:
            try:
                rows = _read_csv_rows(m["csv_path"])
            except Exception:
                return ""
        for row in rows:
            ts = row.get("timestamp_utc", "")
            if ts:
                return ts
        return ""

    sorted_matches = sorted(all_matches, key=_match_date)
    split_idx  = int(len(sorted_matches) * 0.70)
    test_temp  = sorted_matches[split_idx:]

    # Build random k folds
    import random as _random
    rng = _random.Random(seed)
    shuffled = sorted_matches[:]
    rng.shuffle(shuffled)
    n_total   = len(shuffled)
    fold_size = n_total // k
    folds = [shuffled[i * fold_size:(i + 1) * fold_size] for i in range(k)]
    if n_total % k:
        folds[-1].extend(shuffled[k * fold_size:])

    cfg_global  = _load_config()
    min_dur_map = cfg_global.get("min_duration", {})
    registry    = {e[0]: e for e in _STRATEGY_REGISTRY}

    robust_keys  = []
    fragile_keys = []
    cv_details   = {}

    for strat_key, result in individual_results.items():
        entry = registry.get(strat_key)
        if entry is None:
            robust_keys.append(strat_key)
            continue

        params  = result["params"]
        cfg     = _cfg_add_snake_keys({**params, "enabled": True})
        min_dur = min_dur_map.get(strat_key, 1)

        fold_metrics = [
            _crossval_raw_metrics(
                _eval_on_matches_subset(strat_key, entry, cfg, min_dur, fold)
            )
            for fold in folds
        ]
        rois     = [fm["roi"] for fm in fold_metrics]
        mean_roi = sum(rois) / len(rois)
        std_roi  = _math.sqrt(sum((r - mean_roi) ** 2 for r in rois) / len(rois))
        n_pos    = sum(1 for r in rois if r > 0)

        temp_bets    = _eval_on_matches_subset(strat_key, entry, cfg, min_dur, test_temp)
        temp_metrics = _crossval_raw_metrics(temp_bets)

        is_robust = (
            mean_roi >= G_MIN_ROI
            and n_pos >= _math.ceil(k * 0.6)
            and temp_metrics["roi"] >= 0
        )

        cv_details[strat_key] = {
            "folds":     fold_metrics,
            "mean_roi":  round(mean_roi, 1),
            "std_roi":   round(std_roi, 1),
            "n_pos_folds": n_pos,
            "temporal":  temp_metrics,
            "robust":    is_robust,
        }

        fold_str = "  ".join(f"F{i+1}:{fm['roi']:+.0f}%" for i, fm in enumerate(fold_metrics))
        verdict  = "✓ ROBUSTO" if is_robust else "✗ FRÁGIL"
        _log(f"  {strat_key:<28} {fold_str}  "
             f"Mean:{mean_roi:+.1f}%±{std_roi:.1f}  "
             f"Temp:{temp_metrics['roi']:+.1f}%  {verdict}")

        if is_robust:
            robust_keys.append(strat_key)
        else:
            fragile_keys.append(strat_key)
            # Disable in new_strategies
            if strat_key in new_strategies:
                new_strategies[strat_key] = {
                    **new_strategies[strat_key], "enabled": False
                }

    # Remove fragile from individual_results so Phase 3 ignores them
    updated_individual = {k: v for k, v in individual_results.items()
                          if k not in fragile_keys}

    _log(f"\n  Robustas ({len(robust_keys)}): {', '.join(robust_keys)}")
    if fragile_keys:
        _log(f"  Frágiles ({len(fragile_keys)}) → desactivadas: {', '.join(fragile_keys)}")

    return updated_individual, new_strategies, cv_details


# ─────────────────────────────────────────────────────────────────────────────
# PHASE 3 — Portfolio preset generation
# ─────────────────────────────────────────────────────────────────────────────

def phase3_presets(new_strategies: dict, workers: int = 4,
                   dry_run: bool = False) -> dict[str, Path]:
    """
    Write the optimized individual params to a staging config, then run
    optimizer_cli.run() for each of 4 criteria to produce preset files.
    Returns dict { criterion: preset_path }.
    """
    _log("Phase 3 — generando presets de portfolio…")

    # Write staging config (backup first)
    base_cfg = _load_config()
    staging  = {**base_cfg, "strategies": new_strategies}

    bak = CARTERA_BAK_DIR / f"cartera_config_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    if not dry_run:
        shutil.copy2(CARTERA_CFG, bak)
        CARTERA_CFG.write_text(json.dumps(staging, indent=2, ensure_ascii=False),
                                encoding="utf-8")
        _log(f"  Staging config escrito (backup: {bak.name})")
    else:
        _log("  DRY_RUN: staging config NO escrito")

    preset_paths: dict[str, Path] = {}

    for criterion in CRITERIA:
        _log(f"  Ejecutando optimizer_cli — criterio={criterion}…")
        t0 = time.time()
        try:
            result = optimizer_cli.run(
                criterion=criterion,
                bankroll_init=1000.0,
                n_workers=workers,
                out=str(PRESETS_DIR / f"preset_{criterion}_result.json"),
            )
            preset_path = PRESETS_DIR / f"preset_{criterion}_config.json"
            if preset_path.exists():
                preset_paths[criterion] = preset_path
                _log(f"  ✓ {criterion}: {preset_path.name} "
                     f"({time.time()-t0:.0f}s)")
            else:
                _log(f"  ✗ {criterion}: preset no generado")
        except Exception as e:
            _log(f"  ✗ {criterion}: error — {e}")

    if dry_run:
        _log("  DRY_RUN: restaurando config original…")
    return preset_paths


# ─────────────────────────────────────────────────────────────────────────────
# PHASE 4 — Select best preset and apply
# ─────────────────────────────────────────────────────────────────────────────

def _merge_preset_strategies(preset_cfg: dict, base_cfg: dict) -> dict:
    """
    Return merged strategies dict applying preset on top of base config.
    - Preset enables  a strategy → apply all preset params.
    - Preset disables a strategy → only set enabled=False, keep base params.
    - Strategies not in base config → skipped (prevents re-adding obsolete keys).
    """
    merged = dict(base_cfg.get("strategies", {}))
    for k, v in preset_cfg.get("strategies", {}).items():
        if k not in merged:
            continue
        if v.get("enabled"):
            merged[k] = {**merged.get(k, {}),
                         **{pk: pv for pk, pv in v.items() if pv is not None}}
        else:
            merged[k] = {**merged.get(k, {}), "enabled": False}
    return merged


def _eval_preset_real_stats(preset_cfg: dict, base_cfg: dict) -> dict:
    """
    Evaluate a preset using analyze_cartera-equivalent logic.
    Returns real stats (n, wr, roi, pl, ci_low) reflecting what the live system
    would actually produce — not the portfolio optimizer's internal cherry-pick.

    Applies global min_odds floor and market dedup (same as analyze_cartera).
    """
    strategy_configs = _merge_preset_strategies(preset_cfg, base_cfg)
    md = base_cfg.get("min_duration", {})

    all_bets = []
    for (_key, _name, _trigger_fn, _desc, _extract_fn, _win_fn) in _STRATEGY_REGISTRY:
        s_cfg = strategy_configs.get(_key, {})
        if not s_cfg.get("enabled"):
            continue
        min_dur = md.get(_key, 1)
        bets = _analyze_strategy_simple(
            _key, _trigger_fn, _extract_fn, _win_fn,
            _cfg_add_snake_keys(s_cfg), min_dur,
        )
        all_bets.extend(bets)

    # Global min_odds floor (mirrors analyze_cartera).
    # Read from preset_cfg — each preset has its own optimized adjustments.
    global_min_odds = (preset_cfg.get("adjustments") or {}).get("min_odds") or 0
    if global_min_odds > 0:
        all_bets = [b for b in all_bets if (b.get("back_odds") or 0) >= global_min_odds]

    # Market dedup: one bet per market per match, earliest by minuto wins
    # (mirrors analyze_cartera — uses _normalize_mercado for BT/LIVE alignment)
    seen_market: dict = {}
    deduped: list = []
    for b in sorted(all_bets, key=lambda x: x.get("minuto", 0) or 0):
        mid = b.get("match_id", "")
        mkt = _normalize_mercado(b.get("mercado", ""))
        mkey = (mid, mkt)
        if mkey not in seen_market:
            seen_market[mkey] = True
            deduped.append(b)
    all_bets = deduped

    n = len(all_bets)
    if n == 0:
        return {"n": 0, "wr": 0.0, "roi": 0.0, "pl": 0.0, "ci_low": 0.0}
    wins = sum(1 for b in all_bets if b["won"])
    pl   = sum(b["pl"] for b in all_bets)
    ci_l, _ = _wilson_ci(wins, n)
    return {
        "n":      n,
        "wr":     round(wins / n * 100, 1),
        "roi":    round(pl / n * 100, 1),
        "pl":     round(pl, 2),
        "ci_low": ci_l,
    }


def phase4_apply(preset_paths: dict[str, Path],
                 selector: str = DEFAULT_SELECTOR,
                 dry_run: bool = False) -> str | None:
    """
    Compare preset stats, select best, apply to cartera_config.json.
    Returns winning criterion name or None if no valid preset found.
    """
    _log(f"Phase 4 — seleccionando mejor preset (criterio={selector})…")

    base_cfg = json.loads(CARTERA_CFG.read_text(encoding="utf-8"))

    candidates = []
    for crit, path in preset_paths.items():
        try:
            preset_cfg = json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            continue

        _log(f"  Evaluando {crit} con stats reales…")
        rs = _eval_preset_real_stats(preset_cfg, base_cfg)
        n, wr, roi, ci_l, pl = rs["n"], rs["wr"], rs["roi"], rs["ci_low"], rs["pl"]
        _log(f"    {crit}: N={n} WR={wr}% ROI={roi}% IC={ci_l} P/L={pl}")

        if n < MIN_PRESET_N or ci_l < IC95_MIN_LOW:
            _log(f"  SKIP {crit}: N={n} ci_low={ci_l} — no pasa quality gate")
            continue

        score_map = {
            "confident_roi": ci_l * roi / 100,
            "max_wr":        wr,
            "max_roi":       roi,
            "max_pl":        pl,
            "min_dd":        pl,   # no max_dd available in real stats; use pl as proxy
            # Balances statistical quality (ci_low, wr) with sample size (sqrt(N)).
            # Penalizes low-N presets even if they have high ROI.
            # Formula: ci_low × wr × sqrt(N)
            "robust":        ci_l * wr * math.sqrt(max(n, 1)),
        }
        score = score_map.get(selector, ci_l * roi / 100)
        candidates.append({
            "criterion": crit, "score": score,
            "n": n, "wr": wr, "roi": roi, "ci_low": ci_l, "pl": pl,
        })

    if not candidates:
        _log("  No hay presets válidos — cartera_config.json sin cambios")
        return None

    best = max(candidates, key=lambda x: x["score"])
    _log(f"  GANADOR: {best['criterion']} — "
         f"N={best['n']} WR={best['wr']}% ROI={best['roi']}% "
         f"IC=[{best['ci_low']}] P/L={best['pl']}")
    _log(f"  Otros presets: {[c['criterion'] for c in candidates if c != best]}")

    src = preset_paths[best["criterion"]]
    if not dry_run:
        bak = CARTERA_BAK_DIR / f"cartera_config_{datetime.now().strftime('%Y%m%d_%H%M%S')}_final.json"
        shutil.copy2(CARTERA_CFG, bak)

        preset_cfg = json.loads(src.read_text(encoding="utf-8"))
        merged = dict(base_cfg)
        for key in ("flat_stake", "initial_bankroll", "bankroll_mode",
                    "active_preset", "risk_filter", "adjustments"):
            if key in preset_cfg:
                merged[key] = preset_cfg[key]
        merged["strategies"] = _merge_preset_strategies(preset_cfg, base_cfg)
        merged.pop("versions", None)  # legacy key — no longer used
        # Guard: remove any strategy keys that leaked to root level
        for _strat_key in list(merged.keys()):
            if _strat_key in _STRATEGY_REGISTRY_KEYS:
                del merged[_strat_key]

        CARTERA_CFG.write_text(json.dumps(merged, indent=2, ensure_ascii=False),
                               encoding="utf-8")
        _log(f"  ✓ cartera_config.json actualizado desde {src.name}")
        _log(f"    (backup: {bak.name})")
    else:
        _log(f"  DRY_RUN: se aplicaría {src.name}")

    return best["criterion"]


# ─────────────────────────────────────────────────────────────────────────────
# PHASE 5 — Export
# ─────────────────────────────────────────────────────────────────────────────

def phase5_export():
    """
    Generate CSV + XLSX from the current cartera_config.json (final portfolio).
    """
    _log("Phase 5 — exportando resultados…")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Re-run analyze_cartera with final config (clear cache so it reads fresh config)
    csv_reader.clear_analytics_cache()

    data  = csv_reader.analyze_cartera()
    bets  = data.get("bets", [])
    by_st = data.get("by_strategy", {})

    # ── CSV ──────────────────────────────────────────────────────────────────
    import csv as _csv
    from datetime import timedelta as _td

    def _utc_to_local(ts_str: str, offset_hours: int = 1) -> str:
        """Convierte timestamp UTC a hora local (GMT+offset)."""
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
            try:
                dt = datetime.strptime(ts_str, fmt) + _td(hours=offset_hours)
                return dt.strftime("%Y-%m-%d %H:%M:%S")
            except (ValueError, TypeError):
                continue
        return ts_str

    bets_local = []
    for b in bets:
        ts_local = _utc_to_local(b.get("timestamp_utc", ""))
        bets_local.append({**b, "timestamp_utc": ts_local, "fecha": ts_local[:10]})

    csv_path = EXPORTS_DIR / f"bt_results_{ts}.csv"
    fieldnames = ["fecha", "match_id", "match_name", "strategy", "strategy_label", "strategy_desc",
                  "minuto", "mercado", "score_bet", "score_final",
                  "back_odds", "won", "pl", "País", "Liga", "timestamp_utc"]
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        w = _csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        w.writeheader()
        w.writerows(bets_local)
    _log(f"  CSV: {csv_path.name}  ({len(bets)} bets)")

    # ── XLSX ─────────────────────────────────────────────────────────────────
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        # Sheet 1: all bets
        ws_bets = wb.active
        ws_bets.title = "Bets"
        _xlsx_write_rows(ws_bets, fieldnames,
                         [{f: b.get(f, "") for f in fieldnames} for b in bets_local])

        # Sheet 2: strategy summary
        ws_sum = wb.create_sheet("Por Estrategia")
        sum_fields = ["strategy", "bets", "wins", "win_pct", "pl", "roi"]
        sum_rows = []
        for st_key, st_data in sorted(by_st.items()):
            if st_data.get("bets", 0) == 0:
                continue
            ci_l, ci_h = _wilson_ci(st_data.get("wins", 0), st_data.get("bets", 0))
            sum_rows.append({
                "strategy": st_key,
                "bets":     st_data.get("bets", 0),
                "wins":     st_data.get("wins", 0),
                "win_pct":  st_data.get("win_pct", 0),
                "pl":       st_data.get("pl", 0),
                "roi":      st_data.get("roi", 0),
                "ci_low":   ci_l,
                "ci_high":  ci_h,
            })
        _xlsx_write_rows(ws_sum,
                         ["strategy","bets","wins","win_pct","pl","roi","ci_low","ci_high"],
                         sum_rows)

        # Sheet 3: cumulative P/L
        ws_cum = wb.create_sheet("Acumulado")
        ws_cum.append(["#", "strategy", "won", "pl", "cumul_pl"])
        cumul = 0.0
        for i, b in enumerate(bets, 1):
            cumul += b["pl"]
            ws_cum.append([i, b.get("strategy",""), b.get("won",""),
                           b.get("pl",""), round(cumul, 2)])

        # Sheet 4: duplicate bets by market
        ws_dup = wb.create_sheet("Duplicados Mercado")
        from collections import Counter
        dup_counter = Counter(
            (b.get("fecha", ""), b.get("match_name", ""), b.get("mercado", ""))
            for b in bets_local
        )
        dup_rows = sorted(
            [{"fecha": k[0], "match_name": k[1], "mercado": k[2], "count": v}
             for k, v in dup_counter.items() if v > 1],
            key=lambda r: (-r["count"], r["fecha"], r["match_name"])
        )
        _xlsx_write_rows(ws_dup,
                         ["fecha", "match_name", "mercado", "count"],
                         dup_rows)
        # highlight rows with count > 1 in orange
        from openpyxl.styles import PatternFill as _PF
        orange_fill = _PF("solid", fgColor="FF9900")
        for row in ws_dup.iter_rows(min_row=2):
            count_cell = row[3]
            if isinstance(count_cell.value, int) and count_cell.value > 1:
                for cell in row:
                    cell.fill = orange_fill

        # Sheet 5: stats by day
        ws_day = wb.create_sheet("Por Día")
        from collections import defaultdict as _dd
        day_map = _dd(lambda: {"bets": 0, "wins": 0, "pl": 0.0})
        for b in bets_local:
            d = b.get("fecha", "")
            day_map[d]["bets"] += 1
            day_map[d]["wins"] += 1 if b.get("won") else 0
            day_map[d]["pl"]   += b.get("pl", 0)
        day_rows = []
        for d in sorted(day_map):
            dm = day_map[d]
            n, w = dm["bets"], dm["wins"]
            ci_l, ci_h = _wilson_ci(w, n)
            day_rows.append({
                "fecha":   d,
                "bets":    n,
                "wins":    w,
                "win_pct": round(w / n * 100, 1) if n else 0,
                "pl":      round(dm["pl"], 2),
                "roi":     round(dm["pl"] / n * 100, 1) if n else 0,
                "ci_low":  ci_l,
                "ci_high": ci_h,
            })
        _xlsx_write_rows(ws_day,
                         ["fecha", "bets", "wins", "win_pct", "pl", "roi", "ci_low", "ci_high"],
                         day_rows)

        xlsx_path = EXPORTS_DIR / f"bt_results_{ts}.xlsx"
        wb.save(str(xlsx_path))
        _log(f"  XLSX: {xlsx_path.name}")
    except ImportError:
        _log("  WARN: openpyxl no disponible — XLSX no generado. "
             "Instala con: pip install openpyxl")

    return csv_path


# ─────────────────────────────────────────────────────────────────────────────
# PHASE 6 — Monte Carlo risk analysis
# ─────────────────────────────────────────────────────────────────────────────

def phase6_monte_carlo(n_fin: int, n_sims: int = 10_000, seed: int = 42) -> dict:
    """Run Monte Carlo analysis on the final portfolio. Read-only — no config changes."""
    _log(f"Phase 6 — Monte Carlo risk analysis ({n_sims:,} sims, seed={seed})…")
    t0 = time.time()

    csv_reader.clear_analytics_cache()
    data = csv_reader.analyze_cartera()
    bets = data.get("bets", [])
    _log(f"  {len(bets)} bets in final portfolio")

    report = monte_carlo.run_full_analysis(
        bets, n_fin, bankroll_init=500.0, n_sims=n_sims, seed=seed,
    )

    monte_carlo.print_report(report)
    _log(f"\n  Phase 6 done in {time.time()-t0:.1f}s")
    return report


def _xlsx_write_rows(ws, headers: list, rows: list):
    """Write headers + rows to a worksheet with basic formatting."""
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill("solid", fgColor="2F5496")
    ws.append(headers)
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")
    for row in rows:
        ws.append([row.get(h, "") for h in headers])


# ─────────────────────────────────────────────────────────────────────────────
# SAVE / LOAD intermediate results
# ─────────────────────────────────────────────────────────────────────────────

def save_results(individual: dict, approved_strategies: dict,
                  cv_details: dict | None = None, mc_report: dict | None = None):
    data = {
        "timestamp":  datetime.now().isoformat(),
        "individual": individual,
        "approved":   approved_strategies,
        "crossval":   cv_details or {},
        "monte_carlo": mc_report or {},
    }
    RESULTS_FILE.write_text(json.dumps(data, indent=2, ensure_ascii=False),
                             encoding="utf-8")
    _log(f"Resultados guardados en {RESULTS_FILE.name}")


def load_results() -> tuple[dict, dict, dict] | None:
    if not RESULTS_FILE.exists():
        return None
    data = json.loads(RESULTS_FILE.read_text(encoding="utf-8"))
    return data.get("individual", {}), data.get("approved", {}), data.get("crossval", {})


# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Furbo BT optimizer")
    parser.add_argument("--phase",
        choices=["all", "individual", "presets", "apply", "export"],
        default="all")
    parser.add_argument("--strategies", default=None,
        help="Comma-separated list of strategy families to optimize (individual phase only)")
    parser.add_argument("--criterion", default=DEFAULT_SELECTOR,
        help=f"Selector criterion for best preset: {DEFAULT_SELECTOR}")
    parser.add_argument("--workers", type=int, default=4)
    parser.add_argument("--dry-run", action="store_true",
        help="Never write to cartera_config.json")
    parser.add_argument("--no-crossval", action="store_true",
        help="Skip Phase 2.5 cross-validation robustness filter")
    parser.add_argument("--no-mc", action="store_true",
        help="Skip Phase 6 Monte Carlo risk analysis")
    parser.add_argument("--mc-sims", type=int, default=10_000,
        help="Number of Monte Carlo simulations (default: 10000)")
    parser.add_argument("--mc-seed", type=int, default=42,
        help="Random seed for Monte Carlo (default: 42)")
    args = parser.parse_args()

    only = [s.strip() for s in args.strategies.split(",")] if args.strategies else None
    dry  = args.dry_run

    _log("=" * 60)
    _log(f"bt_optimizer.py  phase={args.phase}  workers={args.workers}"
         f"  dry_run={dry}")
    _log("=" * 60)

    # ── Phase 0: always load data first ─────────────────────────────────────
    all_matches, n_fin = phase0_load()

    if args.phase in ("all", "individual"):
        individual = phase1_individual(n_fin, workers=args.workers, only=only)
        new_strats = phase2_build_config(individual)

        # Phase 2.5: cross-validation robustness filter
        cv_details = {}
        if not args.no_crossval:
            individual, new_strats, cv_details = phase2_5_crossval(
                individual, new_strats, all_matches
            )

        save_results(individual, new_strats, cv_details)

        # Print summary table
        phase_label = "1+2+CV" if not args.no_crossval else "1+2"
        _log(f"\n── RESUMEN PHASE {phase_label} ──────────────────────────────────────")
        _log(f"{'Estrategia':<28} {'N':>5} {'WR%':>6} {'ROI%':>7} "
             f"{'P/L':>8} {'IC95':>12} {'CV':>10}")
        _log("-" * 82)
        for fam, r in sorted(individual.items(), key=lambda x: -x[1]["score"]):
            ci  = f"[{r['ci_low']}-{r['ci_high']}]"
            cv  = f"μ{cv_details[fam]['mean_roi']:+.0f}%" if fam in cv_details else ""
            _log(f"{fam:<28} {r['n']:>5} {r['wr']:>6} {r['roi']:>7} "
                 f"{r['pl']:>8.1f} {ci:>12} {cv:>10}")

        if args.phase == "individual":
            return

    elif args.phase in ("presets", "apply"):
        saved = load_results()
        if saved is None:
            _log("ERROR: no hay resultados de phase individual guardados. "
                 "Ejecuta --phase individual primero.")
            sys.exit(1)
        individual, new_strats, cv_details = saved

    if args.phase in ("all", "presets"):
        preset_paths = phase3_presets(new_strats, workers=args.workers, dry_run=dry)
        if not preset_paths:
            _log("No se generaron presets — abortando")
            return

        if args.phase == "presets":
            return

    if args.phase in ("all", "apply"):
        if args.phase == "apply":
            preset_paths = {
                c: PRESETS_DIR / f"preset_{c}_config.json"
                for c in CRITERIA
                if (PRESETS_DIR / f"preset_{c}_config.json").exists()
            }
            if not preset_paths:
                _log(f"ERROR: no hay ficheros preset en {PRESETS_DIR}. "
                     "Ejecuta --phase presets primero.")
                sys.exit(1)
        best = phase4_apply(preset_paths, selector=args.criterion, dry_run=dry)
        if best:
            _log(f"\n✓ Preset aplicado: {best}")

    if args.phase in ("all", "export"):
        phase5_export()

    # Phase 6: Monte Carlo risk analysis (always runs in "all", after export)
    if args.phase == "all" and not args.no_mc:
        mc_report = phase6_monte_carlo(n_fin, n_sims=args.mc_sims, seed=args.mc_seed)
        save_results(individual, new_strats, cv_details, mc_report)

    _log("\n✓ bt_optimizer.py completado")


if __name__ == "__main__":
    main()
